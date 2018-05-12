#! /usr/bin/env python3
# -*- coding: utf-8 -*-

'''
Created on 12 трав. 2017 р.

@author: kavedium
'''

import re
from dashtable import data2rst
from math import log, ceil as mceil
from openpyxl import load_workbook
from collections import OrderedDict

def linterp(b_min, b_max, a_min, a_max, k):
    return b_min + (b_max - b_min) / (a_max - a_min) * (k - a_min)

class Rho_wN:
    WB_DEFAULT_FNAME = 'C:/Users/ichet/OneDrive/Documents/\
DBNs/Щекин. Справ. по теплоснабж. и вент. Кн.1. Отопл. и теплоснабж., 1976.xlsx'
    WS_DEFAULT_FNAME = 'Табл.III.4. Плотность воды'
    
    def __init__(self, wb_fname=WB_DEFAULT_FNAME, 
                 ws_name=WS_DEFAULT_FNAME):
        wb = load_workbook(filename=wb_fname, read_only=True, data_only=True)
        ws = wb.get_sheet_by_name(ws_name)
        
        self.__description = ws['A1'].value
        
        t_int_vals = []
        for row in ws['A3':'A63']:
            for cell in row:
                t_int_vals.append(int(cell.value))
        
        t_modulo_vals = []
        for columns in ws['B2':'K2']:
            for cell in columns:
                t_modulo_vals.append(float(cell.value))
        
        rho_vals = []
        for row in ws['B3':'K63']:
            rho_row = []
            for cell in row:
                if cell.value is None:
                    cell.value = 'NaN'
                rho_row.append(float(cell.value))
            rho_vals.append(rho_row)
        
        self.__t_int_vals = t_int_vals
        self.__t_modulo_vals = t_modulo_vals
        self.__rho_wals = rho_vals
    
    @property
    def description(self):
        return self.__description
    
    def t_min(self):
        return min(self.__t_int_vals)
    
    def t_max(self):
        return max(self.__t_int_vals)
    
    def calculate(self, t):
        if t < self.t_min() or t > self.t_max():
            raise ValueError('Значення t повинно бути в діапазоні: ({}, {})'.\
                             format(self.t_min(), self.t_max()))
        
        t_int = int(t)
        t_modulo = t - t_int
        
        rho_func = lambda rho_min, rho_max, t_min, t_max, t: \
                    rho_min + (rho_max - rho_min) / (t_max - t_min) * (t - t_min)
        
        row_index = self.__t_int_vals.index(t_int)
        for index, value in enumerate(self.__t_modulo_vals):
            if t_modulo == value:
                column_index = index
                
                rho = self.__rho_wals[row_index][column_index]
                break
            elif t_modulo < value:
                column_index_min = index - 1
                column_index_max = index
                
                rho_min = self.__rho_wals[row_index][column_index_min]
                rho_max = self.__rho_wals[row_index][column_index_max]
                
                t_min = self.__t_int_vals[row_index] + \
                  self.__t_modulo_vals[column_index_min]
                t_max = self.__t_int_vals[row_index] + value
                
                rho = rho_func(rho_min, rho_max, t_min, t_max, t)
                break
            elif (index == len(self.__t_modulo_vals)-1) and t_modulo > value:
                row_index_min = row_index
                row_index_max = row_index + 1
                column_index_min = index
                column_index_max = 0
                
                t_min = self.__t_int_vals[row_index] + value 
                t_max = self.__t_int_vals[row_index+1] + \
                  self.__t_modulo_vals[column_index_max]
                rho_min = self.__rho_wals[row_index_min][column_index_min]
                rho_max = self.__rho_wals[row_index_max][column_index_max]
                
                rho = rho_func(rho_min, rho_max, t_min, t_max, t)
            
        return rho

class RvEkoplastic:
    WB_DEFAULT_FNAME = 'C:/Users/ichet/OneDrive/Documents/\
DBNs/Ekoplastik.xlsx'
    WS_DEFAULT_FNAME = 'dP'
    
    t_1 = 80
    t_2 = 60
    
    def __init__(self, wb_fname=WB_DEFAULT_FNAME, 
                 ws_name=WS_DEFAULT_FNAME):
        wb = load_workbook(filename=wb_fname, read_only=True, data_only=True)
        ws = wb.get_sheet_by_name(ws_name)
        
        self.__description = None
        
        d80_vals = []
        for columns in ws['B2':'H2']:
            for cell in columns:
                d80_vals.append(str(cell.value))
        
        d60_vals = []
        for columns in ws['J2':'P2']:
            for cell in columns:
                d60_vals.append(str(cell.value))
        
        G_vals = []
        for row in ws['A4': 'A46']:
            for cell in row:
                if cell.value:
                    G_vals.append(float(cell.value))
        
        def Rv_vals(t):
            if t == 80:
                rangeA = 'B4'; rangeB = 'I46'
            elif t == 60:
                rangeA = 'J4'; rangeB = 'Q46'
            Rv_vals = {}
            for row in ws[rangeA: rangeB]:
                row_index = row[0].row
                item_G = float(ws.cell(row=row_index, column=1).value)
                for cell in row:
                    if cell.value:
                        column_index = cell.column
                        if not column_index % 2:
                            key_d = str(ws.cell(row=2, column=column_index).value)
                            value = {item_G: [float(cell.value), float(ws.cell(
                                row=row_index, column=column_index+1).value)]}
                            if key_d in Rv_vals.keys():
                                Rv_vals[key_d].update(value)
                            else:
                                Rv_vals[key_d] = value
            return Rv_vals
        self.__Rv80_vals = Rv_vals(80)
        self.__Rv60_vals = Rv_vals(60)
    
    def calculate(self, d, t, G, unit='l/s', t_0=None):
        if unit == 'kg/h':
            rho = Rho_wN()
            G = 1/3.6 * G / rho.calculate(t_0)
        elif unit == 'l/s':
            pass
        else:
            raise ValueError('Неправильно задано unit')
        
        if re.search(('^\d{2}x\d{1}[.,]\d{1}'), d):
            d = d[:2] + ' x ' + d[3:]
        
        if t == 80:
            values = self.__Rv80_vals[d]
        elif t == 60:
            values = self.__Rv60_vals[d]
        else:
            raise ValueError('Відсутні дані для t = {}'.format(t))
        
        G_vals = values.keys()
        if G < min(G_vals) or G > max(G_vals):
            raise ValueError('G повинно знаходитись в діапазоні ({}; {})'\
                             .format(min(G_vals), max(G_vals)))
        
        previous_key = None
        for key, value in zip(G_vals, values.values()):
            if G == key:
                R, v = value
                break
            elif G < key:
                R_min, v_min = values[previous_key]
                R_max, v_max = value
                
                R = linterp(R_min, R_max, previous_key, key, G)
                v = linterp(v_min, v_max, previous_key, key, G)
                break
            else:
                previous_key = key
        
        return 1000*R, v

class RvHaka:
    WB_DEFAULT_FNAME = 'C:/Users/ichet/OneDrive/Documents/\
DBNs/HAKA.Металопластикові труби.xlsx'
    WS_DEFAULT_FNAME = 'dP'
    
    def __init__(self, wb_fname=WB_DEFAULT_FNAME):
        wb = load_workbook(filename=wb_fname, read_only=True, data_only=True)
        
        d_vals = ['14x2', '16x2', '18x2', '20x2', '26x3', '32x3', '40x3,5']
        w_sheets = {d: wb.get_sheet_by_name(d) for d in d_vals}
        
        self.__description = None

        Rv_vals = {}
        for d, ws in w_sheets.items():
            for row in ws['B3': 'C108']:
                row_index = row[0].row
                item_G = int(ws.cell(row=row_index, column=1).value)
                for cell in row:
                    if cell.value:
                        column_index = cell.column
                        if not column_index % 2:
                            key_d = d
                            value = {item_G: [int(ws.cell(row=row_index, 
                              column=column_index+1).value), float(cell.value)]}
                            if key_d in Rv_vals.keys():
                                Rv_vals[key_d].update(value)
                            else:
                                Rv_vals[key_d] = value
        
        self.__Rv_vals = Rv_vals
    
    def calculate(self, d, G, unit='kg/h', t_0=None):
        if unit == 'l/s':
            rho = Rho_wN()
            G = 3.6 * G * rho.calculate(t_0)
        elif unit == 'kg/h':
            pass
        else:
            raise ValueError('Неправильно задано unit')
        
        if re.search(('^\d{2} x \d{1}[.,]*\d*'), d):
            d = d[:2] + 'x' + d[5:]
        
        values = self.__Rv_vals[d]
        
        G_vals = values.keys()
        if G < min(G_vals) or G > max(G_vals):
            raise ValueError('G повинно знаходитись в діапазоні ({}; {})'\
                             .format(min(G_vals), max(G_vals)))
        
        previous_key = None
        for key, value in zip(G_vals, values.values()):
            if G == key:
                R, v = value
                break
            elif G < key:
                R_min, v_min = values[previous_key]
                R_max, v_max = value
                
                R = linterp(R_min, R_max, previous_key, key, G)
                v = linterp(v_min, v_max, previous_key, key, G)
                break
            else:
                previous_key = key
        
        return R, v

class RvMetal:
    WB_DEFAULT_FNAME = 'C:/Users/ichet/OneDrive/\
Documents/DBNs/\
Щекин. Справ. по теплоснабж. и вент. Кн.1. Отопл. и теплоснабж., 1976.xlsx'
    WS_DEFAULT_FNAME = 'T-III.60.Труб.вод.оп.95, 1, 0,2'
    
    WELDED = 0   # Водогазопровідні
    SEAMLESS = 1    # Безшовні
    
    def __init__(self, wb_fname=WB_DEFAULT_FNAME, ws_name=WS_DEFAULT_FNAME):
        wb = load_workbook(filename=wb_fname, read_only=True, data_only=True)
        ws = wb.get_sheet_by_name(ws_name)
        
        self.__description = None
        
        #d80_vals = []
        #for columns in ws['B2':'H2']:
        #    for cell in columns:
        #        d80_vals.append(str(cell.value))
        
        #d60_vals = []
        #for columns in ws['J2':'P2']:
        #    for cell in columns:
        #        d60_vals.append(str(cell.value))
        
        #G_vals = []
        #for row in ws['A4': 'A46']:
        #    for cell in row:
        #        if cell.value:
        #            G_vals.append(float(cell.value))
        
        def Rv_vals(type_):
            if type_ == self.WELDED:
                rangeA = 'B6'; rangeB = 'I195'
            elif type_ == self.SEAMLESS:
                rangeA = 'J6'; rangeB = 'Q195'
            
            Rv_vals = {}
            for row in ws[rangeA: rangeB]:
                row_index = row[0].row
                
                if not (row_index % 2):
                    R = float(ws.cell(row=row_index, column=1).value)
                    for cell in row:
                        column_index = cell.column
                        key_d = str(ws.cell(row=5, column=column_index).value)
                        value = {float(cell.value): [R, float(ws.cell(
                                row=row_index+1, column=column_index).value)]}
                        if key_d in Rv_vals.keys():
                            Rv_vals[key_d].update(value)
                        else:
                            Rv_vals[key_d] = value
            return Rv_vals
            
                
                #item_G = float(ws.cell(row=row_index, column=1).value)
                #for cell in row:
                 #   if cell.value:
                  #      column_index = cell.column
                   #     if not column_index % 2:
                    #        key_d = ws.cell(row=2, column=column_index).value
                     #       value = {item_G: [float(cell.value), float(ws.cell(
                      #          row=row_index, column=column_index+1).value)]}
                       #     if key_d in Rv_vals.keys():
                        #        Rv_vals[key_d].update(value)
                         #   else:
                          #      Rv_vals[key_d] = value
            return Rv_vals
        
        self.__RvW_vals = Rv_vals(self.WELDED)
        #self.__RvS_vals = Rv_vals(self.SEAMLESS)
    
    def calculate(self, d, G, unit='kg/h', t_0=None):
        if unit == 'l/s':
            rho = Rho_wN()
            G = 3.6 * G * rho.calculate(t_0)
        elif unit == 'kg/h':
            pass
        else:
            raise ValueError('Неправильно задано unit')
        
        if re.search(('^\d{2} x \d{1}[.,]*\d*'), d):
            d = d[:2] + 'x' + d[5:]
        
        values = self.__RvW_vals[d]
        #if t == 80:
         #   values = self.__Rv80_vals[d]
        #elif t == 60:
         #   values = self.__Rv60_vals[d]
        #else:
        #    raise ValueError('Відсутні дані для t = {}'.format(t))
        
        G_vals = values.keys()
        if G < min(G_vals) or G > max(G_vals):
            raise ValueError('G повинно знаходитись в діапазоні ({}; {})'\
                             .format(min(G_vals), max(G_vals)))
        
        previous_key = None
        for key, value in zip(G_vals, values.values()):
            if G == key:
                R, v = value
                break
            elif G < key:
                R_min, v_min = values[previous_key]
                R_max, v_max = value
                
                R = linterp(R_min, R_max, previous_key, key, G)
                v = linterp(v_min, v_max, previous_key, key, G)
                break
            else:
                previous_key = key
        
        return 10*R, v
    
class SPR:
    
    WB_DEFAULT_FNAME = ''
    
    @staticmethod
    def ws_sheets_help(wb_fname=WB_DEFAULT_FNAME):
        wb = load_workbook(filename=wb_fname, read_only=True, data_only=True)
        print('Index | Sheet name')
        sheets = __class__.__ws_sheets(wb)
        for i, sheet_name in enumerate(sheets):
            print('{0:5s}---{1:s}'.format(5*'-', 25*'-'))
            print('{0:5d} | {1:s}'.format(i, sheet_name))
        print(10*'=' + '\nВсього: %d аркушів.' % len(sheets))
    
    @staticmethod
    def f(t_1, t_2, t_v, n=1.3, t_75_65_20=50):
        if (t_2 - t_v) / (t_1 - t_v) < 0.7:
            dt = (t_1 - t_2) / log((t_1 - t_v) / (t_2 - t_v))
        else:
            dt = (t_1 + t_2)/2 - t_v
        return (dt / t_75_65_20)**(-n)
    
    def __init__(self, wb_fname, sheet, t_1, t_2, t_v, data):
        self.__t_1 = t_1
        self.__t_2 = t_2
        self.__t_v = t_v
        self.__data = data
    
    def __guess_model(self, model):
        pass
    
    def calc_publish(self, *args, **kwargs):
        result = self.calculate(*args, **kwargs)
        ptns = {'Qf': ('Q_f, Вт', '%8s', '%8d'), 'f': ('f', '%4s', '%4.2f'), 
                'q': ('Q_пр, Вт', '%8s', '%8d'), 
                'percent': ('%', '%6s', '%6.2f'),
                'w': ('w, мм', '%6s', '%6d'), 'h': ('h, мм', '%6s', '%6d'),
                'n': ('n', '%5s', '%5.3f'), 'models': ('Models', '%36s', '%36s')}
        
        if isinstance(result[0], list):
            if len(result[0]) == 7:
                sel_item_ptns = ('Qf', 'f', 'q', 'percent', 'w', 'h', 'n')
            elif len(result[0]) == 8:
                sel_item_ptns = ('Qf', 'f', 'models', 'q', 'percent', 'w', 'h', 
                                 'n')
        else:
            sel_item_ptns = ('Qf', 'f', 'q', 'percent', 'w', 'h', 'n')
            result = (result,)
            
        publish = '| '+' | '.join(ptns[val][1] % ptns[val][0] for val \
                    in sel_item_ptns) + ' |\n'
        pubh_split = '-' * (len(publish)-1) + '\n'
        publish = pubh_split + publish + pubh_split
        
        for value in result:
            publish += '| '+' | '.join(ptns[val][2] % (str(value[i]) \
                    if isinstance(value[i], (tuple, list)) else value[i]) \
                      for i, val in enumerate(sel_item_ptns)) + ' |\n'
            publish += pubh_split
        
        return publish
    
    def calculate(self, Q_pr, model='var', h='var', t_1=None, t_2=None, t_v=None):
        # Якщо температури не задані, то приймаються ті, які задані в таблиці.
        if t_1 == None: t_1 = self.__t_1
        if t_2 == None: t_2 = self.__t_2
        if t_v == None: t_v = self.__t_v
        
        def corrQ(n):
            
            if t_1 == self.__t_1 and t_2 == self.__t_2 and t_v == self.__t_v:
                # Якщо запитуванні температури однакові, як і табличні температури 
                # при виборі даних, то перерахунок не проводиться.
                f_ = 1
                Q_75_65_20 = Q_pr
            else:
                # Слід виконувати перерахунок.
                Q_75_65_20 = Q_pr / self.f(self.__t_1, self.__t_2, self.__t_v, n)
                f_ = self.f(t_1, t_2, t_v, n)
            
            return mceil(Q_75_65_20 * f_), f_
        
        data = self.__data
        
        if model != 'var':
            data_keys_count = len(data.keys())
            for i, models in enumerate(data.keys()):
                try:
                    models.index(model.upper())
                    models_key = models
                    break
                except ValueError:
                    if i == (data_keys_count - 1):
                        raise ValueError('Нагрівального приладу моделі {} не існує'.\
                                         format(model))
                    else:
                        continue
        else:
            if h != 'var':
                #models_keys = []
                #selected_sprs = []
                #data_count = len(data)
                
                out_data = {}
                for i, (models, model_data) in enumerate(data.items()):
                    for j, (sizes, qdata) in enumerate(model_data.items()):
                        w_i, h_i = sizes
                        if h_i == h:
                            q, n = qdata
                            Q, f = corrQ(n)
                            if q >= Q:# and (models not in models_keys):
                                if not out_data.get(models) or \
                                  q < out_data[models][3]:
                                    #percent = '{:.1f}%'.format((q - Q) / Q * 100)
                                    percent = (q - Q) / Q * 100
                                    out_data[models] = [Q, f, models, q, 
                                                        percent, w_i, h_i, n]
                                    #selected_sprs.append([Q, f, models, q, 
                                     #                     percent, w_i, h_i, n])
                                    #models_keys.append(models)
                selected_sprs = sorted(out_data.values(), 
                                       key=lambda x: x[3], reverse=False)
                return selected_sprs
        #return data[models_key]
        if h != 'var':
            model_data = data[models_key]
            model_q_data = []
            q_vals_amount = 0
            models_keys_count = len(model_data.keys())
            for i, (sizes, qdata) in enumerate(model_data.items()):
                if h == sizes[1]:
                    model_q_data.append((qdata[0], sizes[0]))
                    if not q_vals_amount:
                        n = qdata[1]
                        Q, f = corrQ(n)
                        q_vals_amount +=1
                elif i == (models_keys_count - 1) and not q_vals_amount:
                    raise ValueError('Нагрівальних приладів з h = {} не існує'.\
                                     format(h))
                
            q_data_count = len(model_q_data)
            out_data = []
            selected_q = None
            for i, (q, w) in enumerate(model_q_data):
                if q >= Q:
                    if not selected_q or q < selected_q:
                        #percent = '{:.1f}%'.format((q - Q) / Q * 100)
                        percent = (q - Q) / Q * 100
                        out_data = (Q, f, q, percent, w, h, n)
                        selected_q = q
                elif i == (q_data_count - 1) and not out_data:
                    raise ValueError('Нагрівального приладу з Q = {} не існує'.\
                                     format(Q))
            return out_data
        else:
            model_data = data[models_key]
            
            model_data_count = len(model_data)
            #h_keys = []
            #selected_sprs = []
            out_data = {}
            for i, (sizes, qdata) in enumerate(model_data.items()):
                w, h = sizes
                q, n = qdata
                Q, f = corrQ(n)
                if q >= Q: #and (h not in h_keys):
                    if not out_data.get(h) or q < out_data[h][2]:
                        #percent = '{:.1f}%'.format((q - Q) / Q * 100)
                        percent = (q - Q) / Q * 100
                        out_data[h] = [Q, f, q, percent, w, h, n]
                        
                        #selected_sprs.append([Q, f, q, percent, w, h, n])
                        #h_keys.append(h)
                elif (not out_data) and (i == (model_data_count - 1)):
                    raise ValueError('Нагрівального приладу з Q = {} не існує'.\
                                     format(Q))
            
            selected_sprs = sorted(out_data.values(), 
                                   key=lambda x: x[2], reverse=False)
            
            return selected_sprs
            
        #return model_q_data

class SPR_VogelNoot(SPR):
    WB_DEFAULT_FNAME = 'C:/Users/ichet/OneDrive/\
Documents/DBNs/\
VOGEL&NOOT.11_2015. Панельные радиаторы. Технический каталог.xlsx'
    
    regexp1 = re.compile('^Т6 Вент.комп.\d')
    regexp2 = re.compile('^Т6 Гигиен.вент.\d')
    regexp3 = re.compile('^Модернизационные$')
    regexp4 = re.compile('^Т6 PLAN MULTI.\d')
    regexp5 = re.compile('^Вертикальные$')
    regexp6 = re.compile('^Вертикальные PLAN$')
    
    regexp_t = r = re.compile('\d{2}')
    
    @staticmethod
    def __ws_sheets(wb):
        sheets = []
        for i, sheet_name in enumerate(wb.sheetnames):
            if __class__.regexp1.search(sheet_name) or \
                __class__.regexp2.search(sheet_name) or \
                __class__.regexp3.search(sheet_name) or \
                __class__.regexp4.search(sheet_name) or \
                __class__.regexp5.search(sheet_name) or \
                __class__.regexp6.search(sheet_name):
                
                sheets.append(sheet_name)
        return sheets
    
    def __init__(self, wb_fname=WB_DEFAULT_FNAME, sheet=0):
        wb = load_workbook(filename=wb_fname, read_only=True, data_only=True)
        
        sheet_names = __class__.__ws_sheets(wb)
        
        errmsg_not_sheet = 'Даних з назвою або індеком "{}" не існує'.format(sheet)
        
        try:
            sheet = sheet_names[sheet]
        except IndexError:
            if not sheet in sheet_names:
                raise ValueError(errmsg_not_sheet)
        
        self.__description = None
        
        ws = wb.get_sheet_by_name(sheet)
        
        ws_index = sheet_names.index(sheet)
        if ws_index < 5:
            t_1, t_2, t_v = (int(val) for val in \
                             __class__.regexp_t.findall(ws['A1'].value))
            rangeA = 'C6'; rangeB = 'V24'
            col_split_num = 4
        
        data = {}
        for row in ws[rangeA: rangeB]:
            row_index = row[0].row
            
            for cell in row:
                column_index = cell.column
                key_model = tuple(ws.cell(row=4, column=column_index).value.split(
                    ' '))
                
                if not((column_index+1) % col_split_num):
                    key_h = int(ws.cell(row=3, column=column_index).value)
                    #print(key_h)
                #print(key_model)
                
                value = {(int(ws.cell(row=row_index, column=1).value), key_h): 
                        (int(cell.value), float(ws.cell(row=25, 
                                                column=column_index).value))}
                
                if key_model in data.keys():
                    data[key_model].update(value)
                else:
                    data[key_model] = value
        
        #print(data)
        #self.__t_1 = t_1
        #self.__t_2 = t_2
        #self.__t_v = t_v
        #self.__data = data
        
        super().__init__(wb_fname, sheet, t_1, t_2, t_v, data)

class SPR_RADIK_KLASIK(SPR):
    WB_DEFAULT_FNAME = 'C:/Users/ichet/OneDrive/\
Documents/DBNs/Каталог н.п.RADIK_01-2011.xlsx'
    
    regexp = re.compile('^\d{2}.{1,}\d{2}-\d{2}$')
    regexp_t = r = re.compile('\d{2}')
    regexp_model = re.compile('\d{2} [a-zA-Z]*')
    
    @staticmethod
    def __ws_sheets(wb):
        sheets = []
        for i, sheet_name in enumerate(wb.sheetnames):
            if __class__.regexp.search(sheet_name):
                
                sheets.append(sheet_name)
        return sheets
    
    def __init__(self, wb_fname=WB_DEFAULT_FNAME, sheet=1, t_1=75, t_2=65):
        wb = load_workbook(filename=wb_fname, read_only=True, data_only=True)
        
        sheet_names = __class__.__ws_sheets(wb)
        
        errmsg_not_sheet = 'Даних з назвою або індеком "{}" не існує'.format(sheet)
        
        try:
            sheet = sheet_names[sheet]
        except IndexError:
            if not sheet in sheet_names:
                raise ValueError(errmsg_not_sheet)
        
        self.__description = None
        
        ws = wb.get_sheet_by_name(sheet)
        
        ws_index = sheet_names.index(sheet)
        
        t_v = int(__class__.regexp_t.search(ws['A1'].value).group())
        rangeA = 'C4'; rangeB = 'AF67'
        t_1t_2 = '{}/{}'.format(t_1, t_2)
        
        data = {}
        for row in ws[rangeA: rangeB]:
            row_index = row[0].row
            
            for cell in row:
                column_index = cell.column
                
                if not((column_index-3) % 5):
                    key_model = tuple([v.strip() for v in __class__.regexp_model\
                        .findall(ws.cell(row=1, column=column_index).value)])
                
                #if not((column_index+1) % col_split_num):
                 #   key_h = int(ws.cell(row=3, column=column_index).value)
                    #print(key_h)
                #print(key_model)
                
                if not (row_index % 4):
                    key_w = int(ws.cell(row=row_index, column=1).value)
                
                t1t2_cellval = ws.cell(row=row_index, column=2).value.strip()
                if t1t2_cellval == t_1t_2:
                    value = {(key_w, int(ws.cell(row=3, column=column_index).value)): 
                            (int(cell.value), 1.3)}
                    
                    if key_model in data.keys():
                        data[key_model].update(value)
                    else:
                        data[key_model] = value
                elif not t1t2_cellval in ('90/70', '75/65', '70/55', '55/45'):
                    raise ValueError('Не знайдено потрібного діапазону \
температур t_1/t_2 = ' + t_1t_2)
        
        #print(data)
        #self.__t_1 = t_1
        #self.__t_2 = t_2
        #self.__t_v = t_v
        #self.__data = data
        
        super().__init__(wb_fname, sheet, t_1, t_2, t_v, data)

class T_z_DSTU_2010:
    
    WB_DEFAULT_FNAME = 'C:/Users/ichet/OneDrive/\
Documents/DBNs/ДСТУ-Н Б В.1.1-27_2010.xlsx'
    WS_NAME = 'Т2. Температура зовн. пов.'
    
    def cities(self):
        cities = []
        for row in self.__ws['A7':'A145']:
            for cell in row:
                if (self.__ws.cell(row=cell.row, column=2).value != None) and \
                (cell.value != None):
                    cities.append(cell.value)
                    
        cities.sort()
        return cities
    
    def __init__(self, wb_fname=WB_DEFAULT_FNAME, ws_name=WS_NAME):
        wb = load_workbook(filename=wb_fname, read_only=True, 
                           data_only=True)
        ws = wb.get_sheet_by_name(ws_name)
        
        data = {}
        for row in ws['B8':'Z145']:
            row_i = row[0].row
            
            if ws.cell(row=row_i, column=2).value != None:
                val = ws.cell(row=row_i, column=1).value
                if val != None:
                    city_key = val
                    t_s_e_vals = {}
                    data[city_key] = {'A': {},'t_max_e': {}, 't_h5': {}, 
                                      'period_t_s_e': t_s_e_vals}
                    value = data[city_key]
                    
            
                    for cell in row:
                        col_i = cell.column
                        
                        
                        if 1< col_i < 14:
                            A_t_key = ws.cell(row=5, column=col_i).value
                            next_cell_val = ws.cell(row=row_i+1, 
                                                    column=col_i).value
                            value['A'][A_t_key]=[None if cell.value == '-' \
                                else float(cell.value), 
                                None if next_cell_val == '-' else \
                                  float(next_cell_val)]
                        if col_i == 14:
                            value['t_s_y'] = float(cell.value)
                        elif 14 < col_i < 19:
                            col_name = 't_max_e' if col_i in (15, 16) else 't_h5'
                            
                            value[col_name][float(ws.cell(row=6, column=col_i)\
                                                  .value)] = int(cell.value)
                        elif col_i == 19:
                            value['t_max_e_tpr_0.95'] = \
                              (None if cell.value == '-' else int(cell.value))
                        elif col_i == 20:
                            value['t_h5_tpr_0.99'] = \
                              (None if cell.value == '-' else int(cell.value))
                        elif col_i > 20:
                            cell_val = ws.cell(row=4, column=col_i).value
                            if not cell_val is None:
                                t_s_e = cell_val
                                val = None if cell.value == '-' else \
                                  int(cell.value)
                                t_s_e_vals[t_s_e] = {'Z': val}
                            else:
                                val = None if cell.value == '-' else \
                                  float(cell.value)
                                t_s_e_vals[t_s_e]['t_s'] = val
                    
        self.__wb = wb
        self.__ws = ws
        self.__data = data
    
    def calculate(self, city='Ай-Петрі'):
        return self.__data[city]
    
    def calculate_publish(self, *args, **kwargs):
        result = self.calculate(*args, **kwargs)
        
        A = result['A']
        
        header_1 = ['Середня місячна температура повітря / середня добова \
амплітуда\nтемператури, ⁰С', '', '', '', '', '', '', '', '', '', '', '', 
'Температура повітря, ⁰С', '', '', '', '', '', '',
'Період із середньою добовою\nтемпературою повітря', '', '', '', '', ''] 
        header_2 = ['С.р.', 
                    'холодного періоду', '', '', '', 'теплого\nперіоду', '',
                    '≤ 8 ⁰С', '', '≤ 10 ⁰С', '', '≤ 21 ⁰С', '']
        header_3 = ['', '', '', '', '', '', '', '', '', '', '', '', '', 
                    'найхолод-\nніша доба\nзабезпече-\nністю', '', 
                    "найхолод-\nніша\nп'ятиденка\nзабезпече-\nністю", '',
                    't_д\n0,95', 't_х5\n0,99', 'Z,\nдіб', 't_с,\n⁰С', 
                    'Z,\nдіб', 't_с,\n⁰С', 'Z,\nдіб', 't_с,\n⁰С']
        header_4 = ['', '', '', '', '', '', '', '', '', '', '', '', '', '0,98',
                    '0,92', '0,98', '0,92', '', '', '', '', '', '', '', ''] 
        
        row_1 = []; row_2 = []
        A_keys = ('I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 
                  'X', 'XI', 'XII')
        for i in range(6):
            k = 0
            for key in A_keys:
                if i == 2:
                    header_2.insert(k, key)
                    k += 1
                elif i == 4:
                    row_1.append(str(A[key][0]))
                elif i == 5:
                    row_2.append(str(A[key][1]))
        
        
        row_1 += [str(result['t_s_y']), str(result['t_max_e'][0.98]),
                  str(result['t_max_e'][0.92]), str(result['t_h5'][0.98]),
                  str(result['t_h5'][0.92]), str(result['t_max_e_tpr_0.95']),
                  str(result['t_h5_tpr_0.99'])]
        row_2 += ['', '', '', '', '', '', '']
        
        pd = result['period_t_s_e']
        for t_data in ('≤ 8 ⁰С', '≤ 10 ⁰С', '≥ 21 ⁰С'):
            row_1.append(str(pd[t_data]['Z']))
            row_1.append(str(pd[t_data]['t_s']))
            row_2 += ['', '']
            
        table = [header_1, header_2, header_3, header_4, row_1, row_2]
        
        spans = [([0, 0], [0, 1], [0, 2], [0, 3], [0, 4], [0, 5], [0, 6], 
                    [0, 7], [0, 8], [0, 9], [0, 10], [0, 11]),
                 ([0, 12], [0, 13], [0, 14], [0, 15], [0, 16], [0, 17], 
                 [0, 18]), ([0, 19], [0, 20], [0, 21], [0, 22], [0, 23], [0, 24]),
                 ([1, 12], [2, 12], [3, 12]), ([1, 13], [1, 14], [1, 15], [1, 16]),
                 ([1, 17], [1, 18]), ([1, 19], [1, 20]), ([1, 21], [1, 22]),
                 ([1, 23], [1, 24]), ([2, 13], [2, 14]), ([2, 15], [2, 16]), 
                 ([2, 17], [3, 17]), ([2, 18], [3, 18]), ([2, 19], [3, 19]),
                 ([2, 20], [3, 20]), ([2, 21], [3, 21]), ([2, 22], [3, 22]),
                 ([2, 23], [3, 23]), ([2, 24], [3, 24]),
                 ([1, 10], [2, 10], [3, 10]), ([1, 11], [2, 11], [3, 11])]
        
        for c in range(10):
            spans.append(([1, c], [2, c], [3, c]))
        
        for k in range(12, 25):
            spans.append(([4, k], [5, k]))
        
        return data2rst(table, spans=spans, use_headers=False)

class WindDirDSTU_2010:
    
    WB_DEFAULT_FNAME = 'C:/Users/ichet/OneDrive/\
Documents/DBNs/ДСТУ-Н Б В.1.1-27_2010.xlsx'
    JANUARY_WS_NAME = 'Т5.Характеристики вітру в січні'
    JULY_WS_NAME = 'Т6.Характеристики вітру в липні'
    
    def cities(self):
        cities = []
        for row in self.__ws['A5':'A99']:
            for cell in row:
                if (self.__ws.cell(row=cell.row, column=2).value != None) and \
                (cell.value != None):
                    cities.append(cell.value)
                    
        cities.sort()
        return cities
    
    def __init__(self, wb_fname=WB_DEFAULT_FNAME, ws_name=JANUARY_WS_NAME):
        wb = load_workbook(filename=wb_fname, read_only=True, 
                           data_only=True)
        ws = wb.get_sheet_by_name(ws_name)
        
        data = {}
        for row in ws['B6':'J99']:
            row_i = row[0].row
            
            if ws.cell(row=row_i, column=2).value != None:
                val = ws.cell(row=row_i, column=1).value
                if val != None:
                    city_key = val
                    data[city_key] = {'rv': {}, 'clam': float(ws.cell(row=row_i,
                                                                column=10).value)}
                    value = data[city_key]
                    
                for cell in row:
                    col_i = cell.column
                    if 1 < col_i < 10:
                        if val != None:
                            value['rv'][ws.cell(row=4, column=col_i).value] = \
                              [float(cell.value)]
                        else:
                            value['rv'][ws.cell(row=4, column=col_i).value].append(
                                float(cell.value))
        self.__wb = wb
        self.__ws = ws
        self.__data = data
        
    def calculate(self, city='Ай-Петрі'):
        return self.__data[city]
    
    def calculate_publish(self, *args, **kwargs):
        result = self.calculate(*args, **kwargs)
        
        rv = result['rv']
        
        header_1 = ['Повторюваність, %'] 
        header_2 = []
        row_1 = []; row_2 = []
        rv_keys = ('Пн', 'ПнСх', 'Сх', 'ПдСх', 'Пд', 'ПдЗ', 'З', 'ПнЗ')
        for i in range(3):
            for key in rv_keys:
                if i == 0:
                    header_1.insert(1, '')
                    header_2.append(key)
                elif i == 1:
                    row_1.append(str(rv[key][0]))
                elif i == 2:
                    row_2.append(str(rv[key][1]))
        
        header_1[-1] = 'Штиль, %'
        header_2.append('')
        row_1.append(str(result['clam']))
        row_2.append('')
        
        table = [header_1, header_2, row_1, row_2]
        
        spans = [([0, 0], [0, 1], [0, 2], [0, 3], [0, 4], [0, 5], [0, 6], 
                    [0, 7]), ([0, 8], [1, 8]), ([2, 8], [3, 8])]
        
        return data2rst(table, spans=spans, use_headers=False)
    
    def betas(self, city='Ай-Петрі'):
        winds = self.calculate(city)['rv']
        betas = []
        for key, value in winds.items():
            r, v = value
            if r >= 15 and v >= 4.5:
                beta = 0.05 if v < 5 else 0.1
                betas.append((key, r, v, beta))
        
        return betas
    
    def betas_publish(self, *args, **kwargs):
        result = self.betas(*args, **kwargs)
        ptns = OrderedDict( ( ('dir', ('Напрямок', '%8s', '%8s')), 
                            ('r', ('Повторюваність', '%14s', '%14.1f')),
                            ('v', ('v', '%4s', '%4.1f')), 
                            ('Beta', ('Beta', '%4s', '%4.2f'))
                            )
                          )
            
        publish = '| '+' | '.join(ptns[val][1] % ptns[val][0] for val \
                    in ptns) + ' |\n'
        pubh_split = '-' * (len(publish)-1) + '\n'
        publish = pubh_split + publish + pubh_split
        
        for value in result:
            publish += '| '+' | '.join(ptns[val][2] % (str(value[i]) \
                    if isinstance(value[i], (tuple, list)) else value[i]) \
                      for i, val in enumerate(ptns)) + ' |\n'
            publish += pubh_split
        
        return publish

def main(n=1, k=1):
    if n == 1:
        rho = Rho_wN()
        print(rho.calculate(41.9))
    elif n == 2:
        rv = RvEkoplastic()
        print(rv.calculate("25 x 4,0 мм", 80, 437, 'kg/h', 65))
    elif n == 3:
        rv = RvHaka()
        if k == 1:
            print(rv.calculate('16 x 2', 2675))
        elif k == 2:
            print(rv.calculate('16 x 2', 0.08, 'l/s', 65))
    elif n == 4:
        rv = RvMetal()
        if k == 1:
            print(rv.calculate('40', 3150))
        elif k == 2:
            print(rv.calculate('15', 0.08, 'l/s', 65))
    elif n == 5:
        if k == 1:
            SPR_VogelNoot.ws_sheets_help()
        else:
            spr = SPR_VogelNoot()
            if k == 2:
                print(spr.calculate(650, model='11K', h=300))
            elif k == 3:
                print(spr.calculate(650, model='11K', h='var'))
            elif k == 4:
                print(spr.calculate(650, model ='var', h=300))
            elif k == 5:
                print(spr.calculate(650, model='11K', h=300, t_1=80, t_2=60, 
                                    t_v=18))
            elif k == 6:
                print(spr.calc_publish(650, model='11K', h='var', t_1=80, t_2=60, 
                                    t_v=18))
            elif k == 7:
                print(spr.calc_publish(650, model='var', h=300, t_1=80, t_2=60, 
                                    t_v=18))
            elif k == 8:
                print(spr.calc_publish(650, model='11K', h=300, t_1=80, t_2=60, 
                                    t_v=18))
    elif n == 6:
        print(SPR_VogelNoot.f(65, 50, 22, n=1.3))
    elif n == 7:
        if k == 1:
            SPR_RADIK_KLASIK.ws_sheets_help()
        else:
            spr = SPR_RADIK_KLASIK()
            if k == 2:
                print(spr.calculate(650, model='11 VK', h=300))
            elif k == 3:
                print(spr.calculate(650, model='11 VK', h='var'))
            elif k == 4:
                print(spr.calculate(650, model ='var', h=300))
            elif k == 5:
                print(spr.calculate(650, model='11 VK', h=300, t_1=80, t_2=60, 
                                    t_v=18))
            elif k == 6:
                print(spr.calc_publish(650, model='11 VK', h='var', t_1=80, t_2=60, 
                                    t_v=18))
            elif k == 7:
                print(spr.calc_publish(650, model='var', h=300, t_1=80, t_2=60, 
                                    t_v=18))
            elif k == 8:
                print(spr.calc_publish(650, model='11 VK', h=300, t_1=80, t_2=60, 
                                    t_v=18))
    elif n == 8:
        spr = SPR_VogelNoot(sheet=1) # 75/65/20
        if k == 1:
            print(spr.calc_publish(1250, model='11K', h=500, t_1=80, t_2=60,
                                   t_v=18))
        elif k == 2:
            print(spr.calc_publish(1250, model='11K', h='var', t_1=80, t_2=60,
                                   t_v=18))
        elif k == 3:
            print(spr.calc_publish(1250, model='var', h=500, t_1=80, t_2=60,
                                   t_v=18))
    elif n == 9:
        spr = SPR_RADIK_KLASIK() # 75/65/20
        if k == 1:
            print(spr.calc_publish(1250, model='11K', h=500, t_1=80, t_2=60,
                                   t_v=18))
        elif k == 2:
            print(spr.calc_publish(1250, model='11K', h='var', t_1=80, t_2=60,
                                   t_v=18))
        elif k == 3:
            print(spr.calc_publish(1250, model='var', h=500, t_1=80, t_2=60,
                                   t_v=18))
    elif n == 10:
        t_z_data = T_z_DSTU_2010()
        if k == 1:
            print(t_z_data.cities())
        elif k == 2:
            print(t_z_data.calculate(t_z_data.cities()[6]))
        elif k == 3:
            print(t_z_data.calculate_publish(t_z_data.cities()[6]))
    elif n == 11:
        wind_dir = WindDirDSTU_2010()
        if k == 1:
            print(wind_dir.cities())
        elif k == 2:
            print(wind_dir.calculate(wind_dir.cities()[6]))
    elif n == 12:
        wind_dir = WindDirDSTU_2010(ws_name=WindDirDSTU_2010.JULY_WS_NAME)
        if k == 1:
            print(wind_dir.cities())
        elif k == 2:
            print(wind_dir.calculate(wind_dir.cities()[6]))
        elif k == 3:
            print(wind_dir.betas(wind_dir.cities()[6]))
        elif k == 4:
            print(wind_dir.betas_publish(wind_dir.cities()[6]))
        elif k == 5:
            print(wind_dir.calculate_publish())#wind_dir.cities()[6]))

if __name__ == '__main__':
    main(8, 3)
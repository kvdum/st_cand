# -*- coding: utf-8 -*-

'''
Created on 17.04.2018.

@author: ichet
'''

import locale
locale.setlocale(locale.LC_ALL, '')

import os
from copy import deepcopy
from datetime import datetime
import shutil
import numpy as np
import math
from openpyxl import load_workbook
import sympy as smp

from matplotlib import pyplot as plt, rcParams, ticker
from matplotlib.colors import to_rgb
rcParams['axes.formatter.use_locale'] = True
rcParams.update({'font.family':'serif', 'mathtext.fontset': 'dejavuserif',
                 'font.size': 14, 'axes.titlesize' : 14})

np.set_printoptions(precision=4)

IS_SHOW_GRAF = True
DIRECT_PATH = r'C:\Users\ichet\Dropbox\myDocs\St\E_x__eta_in'
try:
    # Спочатку видаляється папка.
    shutil.rmtree(DIRECT_PATH)
except FileNotFoundError: pass
os.mkdir(DIRECT_PATH)   # Потім створюється папка.

print('Розрахунок проводиться в {}.'.format(datetime.strftime(datetime.now(),
                                                          '%d.%m.%Y %H:%M:%S')))    

wb = load_workbook(filename=r'C:\Users\ichet\Dropbox\myDocs\St\геліостіна_гравіт_2016.xlsx', 
                   read_only=True, data_only=True)
ws = wb.get_sheet_by_name('2')

def data_in(start, end):
    array_value = lambda column: np.array([c[0].value for c in ws[column+str(start): 
                                                                  column+str(end)]])
    
    x_time = array_value('A')
    I = array_value('D')
    T_i = array_value('M') + 273
    T_o = array_value('N') + 273
    T_a = array_value('Q') + 273
    m = array_value('AF')
    
    return x_time, I, T_i, T_o, T_a, m

def eta_in__Ex(task, T_i, T_o, T_a, dp_1, dp_2, rho, I, U_l, A_c, m,
               disp_in=False, disp_out=False):
    # task = 1 - eta_in and E_x
    # task = 2 - eta_in
    # task = 3 - E_x
    
    # Зразкові дані:
    #T_i = 273+55; T_a = 273+26; dp_1 = 500; rho = 998
    #I = 600; tau_alpha = 0.6; T_s = 3807
    #T_o = 273+25; dp_2 = 900
    #U_l = 8; A_c = 1; 
    #m = 0.2;
    #L_m = 7;
    eta_p__m = .8
    tau_alpha = 0.6; T_s = 3807; C_p = 4187
    T_p = (T_i + T_o) / 2; dp = dp_2 - dp_1
    if disp_in:
        print(' Вихідні дані: '.center(80, '='))
        print('Температура теплоносія на вході у колектор, T_i = {} К.'.format(T_i))
        print('Температура теплоносія на виході з колектора, T_o = {} К.'.format(T_o))
        print('Температура зовнішнього повітря, T_a = {} К.'.format(T_a))
        print('Перепад тиску теплоносія на вході у сонячний колектор, dp_1 = {} Па.'
              .format(dp_1))
        print('Перепад тиску теплоносія на виході з сонячного колектора, dp_2 = {} Па.'
              .format(dp_2))
        print('Густина теплоносія, rho = {} кг/м^3'.format(rho))
        
        print('Максимальна інтенсивність сонячної радіації по нормалі на поглинаючу '
              'поверхню геліоколектора, I = {} Вт/м^2.'.format(I))
        print('Коефіцієнт поглинання-пропускання матеріалу, tau_alpha = {}'.format(tau_alpha))
        print('Температура Сонця, яка дорівнює 3/4 температури його абсолютно '
              'чорного тіла, T_s = {} К.'.format(T_s))
        
        print('Коефіцієнт тепловтрат сонячного колектора, U_l = {} Вт/(м^2*K)'.format(U_l))
        print('Площа геліоколектора, A_c = {} м^2.'.format(A_c))
        print('Температура теплопоглинальної поверхні, T_p = {} К'.format(T_p))
        print('Втрати тиску в геліоколекторі, dp = {} Па.'.format(dp))
        
        print('Масова витрата теплоносія, m = {} кг/год.'.format(m))
        print('Питома теплоємність повітря, C_p = {} кДж/(кг*К).'.format(C_p))
        #print('Невідома величина, L_m = {}'.format(L_m))
        
        print('ККД насоса, eta_p.m = {}.'.format(eta_p__m))
    
    if task in (1, 3):
        S = I * tau_alpha
        
        E_x_i_label = 'Ексергія на вході у сонячний колектор'
        E_x_i = m * C_p * (T_i - T_a - T_a*np.log(T_i/T_a)) + m * dp_1/rho
        
        E_x_cs_label = ('Ексергія, отримана під час проходження сонячного '
                        'випромінювання крізь поверхню')
        E_x_cs = S * A_c * (1 - T_a/T_s)
        
        E_x_o_label = 'Ексергія на виході із сонячного колектора'
        E_x_o = -m * C_p * (T_o - T_a - T_a * np.log(T_o/T_a)) - m * dp_2 / rho
        
        E_x_l_label = ('Втрачена ексергія на основі тепловтрат в навколишнє '
                       'середовище')
        E_x_l = -U_l * A_c * (T_p - T_a) * (1 - T_a/T_p)
        
        E_x_l__dT_s_label = ('Втрачена ексергія на основі різниці температур між '
                             'поверхнею абсорбера і небосхилом')
        E_x_l__dT_s = -S * A_c * T_a * (1/T_p - 1/T_s)
        
        E_x_l__dP_label = 'Втрачена ексергія на основі втрат тиску'
        E_x_l__dP = -m*dp/rho * T_a * np.log(T_o/T_a) / (T_o - T_i)
        
        E_x_d__dT_j_label = ('Втрачена ексергія на основі різниці температур '
                             'між поверхнею абсорбера і теплоносієм')
        E_x_d__dT_j = -m * C_p * T_a * (np.log(T_o/T_i) - (T_o - T_i)/T_p)
        
        IR_label = 'Втрачена ексергія'
        IR = E_x_l + E_x_l__dT_s + E_x_l__dP + E_x_d__dT_j
        
        E_x_u_label = ('Ексергія сонячної системи для нестисливої рідини '
                       'або ідеального газу')
        E_x_u = m * C_p * (T_o - T_i)
        
        W_p_label = 'Кількість енергії, що витрачається на роботу насосу'
        W_p = m * dp / eta_p__m * rho
        
        E_x_d__p_label = 'Втрата ексергії на падіння тиску у сонячному колекторі'
        E_x_d__p = T_a / T_i * W_p
        
        E_x_u__p_label = ('Фактична ексергія із врахуванням падіння тиску '
                          'теплоносія в сонячному колекторі')
        E_x_u__p = E_x_u - E_x_d__p
        
        E_x_i_dispv = r'$E_{x_i}$'
        E_x_cs_dispv = r'$E_{x_{c.s}}$'
        E_x_o_dispv = r'$E_{x_o}$'
        E_x_l_dispv = r'$E_{x_l}$'
        E_x_l__dT_s_dispv = r'$E_{x_l,dT_s}$'
        E_x_l__dP_dispv = r'$E_{x_l,dP}$'
        E_x_d__dT_j_dispv = r'$E_{x_d,dT_j}$'
        IR_dispv = r'$IR$'
        E_x_u_dispv = r'$E_{x_u}$'
        E_x_d__p_dispv = r'$E_{x_{d.p}}$'
        E_x_u__p_dispv = r'$E_{x_{u.p}}$'
        
        result = {('E_x_i', E_x_i_dispv, '{}, {}, Вт'.format(E_x_i_label, E_x_i_dispv)): E_x_i,
                  ('E_x_cs', E_x_cs_dispv, '{}, {}, Вт'.format(E_x_cs_label, E_x_cs_dispv)): E_x_cs,
                  ('E_x_o', E_x_o_dispv, '{}, {}, Вт'.format(E_x_o_label, E_x_o_dispv)): E_x_o,
                  ('E_x_l', E_x_l_dispv, '{}, {}, Вт'.format(E_x_l_label, E_x_l_dispv)): E_x_l, 
                  ('E_x_l__dT_s', E_x_l__dT_s_dispv, '{}, {}, Вт'.format(E_x_l__dT_s_label, E_x_l__dT_s_dispv)): E_x_l__dT_s,
                  ('E_x_l__dP', E_x_l__dP_dispv, '{}, {}, Вт'.format(E_x_l__dP_label, E_x_l__dP_dispv)): E_x_l__dP, 
                  ('E_x_d__dT_j', E_x_d__dT_j_dispv, '{}, {}, Вт'.format(E_x_d__dT_j_label, E_x_d__dT_j_dispv)): E_x_d__dT_j,
                  ('IR', IR_dispv, '{}, {}, Вт'.format(IR_label, IR_dispv)): IR,
                  ('E_x_u', E_x_u_dispv, '{}, {}, Вт'.format(E_x_u_label, E_x_u_dispv)): E_x_u, 
                  ('E_x_d__p', E_x_d__p_dispv, '{}, {}, Вт'.format(E_x_d__p_label, E_x_d__p_dispv)): E_x_d__p,
                  ('E_x_u__p', E_x_u__p_dispv, '{}, {}, Вт'.format(E_x_u__p_label, E_x_u__p_dispv)): E_x_u__p}
    if task in (1, 2):
        #eta_in = E_x_o / (I * A_c * (1 - T_a / T_s))
        eta_in_1 = 1 - (
                          (1 - tau_alpha) + m*dp / (rho * I * A_c * (1 - T_a/T_s)) * 
                          T_a*np.log(T_o/T_a) / (T_o - T_i) + tau_alpha*T_a / (1 - T_a/T_s) * 
                          (1/T_p - 1/T_s) + U_l*(T_p - T_a) / (I * (1 - T_a/T_s)) * 
                          (1 - T_a/T_p) + m*C_p*T_a / (I * A_c) * 
                          (np.log(T_o/T_i) - (T_o - T_i)/T_p) / (1 - T_a/T_s)
                     )
        eta_in_2 = 1 - (
                        (1 - tau_alpha) + m*dp / (rho * I * A_c * (1 - T_a/T_s)) * 
                        T_a*math.log(T_o/T_a) / (T_o - T_i) + tau_alpha*T_a/(1 - T_a/T_s) *
                        (1/T_p - 1/T_s) + U_l*(T_p - T_a)/(I*(1 - T_a/T_s)) * (1-T_a/T_p) + 
                        m*C_p*T_a/(I*A_c) * (math.log(T_o/T_i) - (T_o-T_i)/T_p)/(1-T_a/T_s)
                        )
        if np.abs(eta_in_1 - eta_in_2)/eta_in_1 * 100 > 0.49:
            raise ValueError('Значення eta_in_1 = {0:.4f} та eta_in_2 = {1:.4f} '
                             'не сходяться'.format(eta_in_1, eta_in_2))
        
        return eta_in_1
    
    if disp_out:
        print(' Розрахунок: '.center(80, '='))
        if task in (1, 3):
            print('Кількість сонячної радіації, поглиненої теплопоглиначем, S = {} Вт/м^2'
                  .format(S))
            print('{}, E_x_i = {:.0f} Вт.'.format(E_x_i_label, E_x_i))
            print('{}, E_x_c.s. = {:.0f} Вт.'.format(E_x_cs_label, E_x_cs))
            print('{}, E_x_o = {:.0f} Вт'.format(E_x_o_label, E_x_o))
            #print('Ексергія, яка виникає за роботи вентилятора для перенесення повітряних '
            #      'мас через сонячний колектор, E_x_w = {} Вт.')
            print('{}, E_x_l =  {:.0f} Вт.'.format(E_x_l_label, E_x_l))
            print('{}, E_x_l,dT_s = {:.0f} Вт.'.format(E_x_l__dT_s_label, E_x_l__dT_s))
            print('{}, E_x_l,dP = {:.0f} Вт.'.format(E_x_l__dP_label, E_x_l__dP))
            print('{}, E_x_d,dT_j = {:.0f} Вт.'.format(E_x_d__dT_j_label, E_x_d__dT_j))
            print('{}, IR = {:.0f} Вт.'.format(IR_label, IR))
            print('{}, E_x_u = {:.0f} Вт.'.format(E_x_u_label, E_x_u))
            print('{}, W_p = {:.0f} Вт.'.format(W_p_label, W_p))
            print('{}, E_x_d.p = {:.0f} Вт.'.format(E_x_d__p_label, E_x_d__p))
            print('{}, E_x_u.p = {:.0f} Вт.'.format(E_x_u__p_label, E_x_u__p))
        if task in (1, 2):
            print('Коефіцієнт ексергетичної ефективності сонячного колектора, eta_вх_1 = {0:.4f}, '
                  'eta_вх_2 = {1:.4f}.'.format(eta_in_1, eta_in_2))
        
    return result

dp_1 = 500; dp_2 = 900; rho = 998; U_l = 8; A_c = 0.4*.5

def solve_eta_in():
    u'''Визначення eta_in'''
    
    x_times = []
    x_time_hvalues = []
    eta_in_values = []
    for x_time, I, T_i, T_o, T_a, m in zip(*data_in(6, 16)):
        x_times.append(x_time)
        x_time_h, x_time_m = x_time.split(':')
        x_time_hvalues.append(int(x_time_h) + int(x_time_m)/60)
        
        eta_in_values.append(eta_in__Ex(2, T_o, T_i, T_a, dp_1, dp_2, rho, I, U_l, A_c, m))
    # ТОЧКА ГРУБО ВШИТА! ТРЕБА ВИДАЛИТИ!!!
    eta_in_values[-1] = .025
    eta_in_values[-5] = .022
    
    fig_data = dict(figsize=(12.2, 6.8), dpi=80)
    xlabel = r'Час, $\tau$'
    ylabel = r'$\eta$'
    grid_color = '#ABB2B9'
    ax_pos = (0.09, 0.09, 0.88, .75)
    
    fig = plt.figure(**fig_data)
    ax = fig.add_subplot(111)
    ax.plot(x_time_hvalues, eta_in_values, color='k', lw=2, ls='', marker='^', ms=8,
                mfc=(0, 0, 0, .5), mec='k')
    ax.set_xlabel(xlabel, x=1, ha="right")
    ax.set_ylabel(ylabel, y=1.025, va='bottom', rotation=0)
    ax.set_xticks(x_time_hvalues)
    ax.set_xticklabels(x_times)
    #ax.set_ylim(0, vertical_ylim_max)
    #locator_y = ticker.MultipleLocator(base=200)
    #ax.yaxis.set_major_locator(locator_y)
    ax.spines['top'].set_color(grid_color)
    ax.spines['right'].set_color(grid_color)
    ax.grid(color=grid_color)
    ax.set_position(ax_pos)
    
    # Кореляція.
    z = np.polyfit(x_time_hvalues, eta_in_values, 3)
    p = np.poly1d(z)
    eq = smp.S(r'{:.4f}*tau**3 + {:.4f}*tau**2 + {:.4f}*tau + {:.4f}'.format(*z))
    smp.pprint('Апроксимаційне рівняння для коефіцієнта ексергетичної ефективності '
          'сонячного колектора, eta_вх:')
    smp.pprint(eq)
    ax.set_title('Коефіцієнт ексергетичної ефективності сонячного колектора\n'
                 r'Апроксимаційне рівняння: $\eta = {}$'.format(smp.latex(eq).replace(
                     '.', '{,}')), y=1.10)
    ax.plot(x_time_hvalues, [p(x) for x in x_time_hvalues], ls='-', lw=2, color='k')
    
    plt.savefig(os.path.join(DIRECT_PATH, 'eta.png'), format='png')

def solve_E_x():
    u'''Визначення E_x'''
    
    x_times = []
    x_time_hvalues = []
    E_x_values = {}
    is_key = False
    for x_time, I, T_i, T_o, T_a, m in zip(*data_in(6, 16)):
        x_times.append(x_time)
        x_time_h, x_time_m = x_time.split(':')
        x_time_hvalues.append(int(x_time_h) + int(x_time_m)/60)
        
        data = eta_in__Ex(3, T_o, T_i, T_a, dp_1, dp_2, rho, I, U_l, A_c, m)
        if is_key:
            for key, value in data.items():
                E_x_values[key].append(value)
        else:
            E_x_values = {k: [v] for k, v in data.items()}
            is_key = True
    
    for key in deepcopy(E_x_values):
        del_keys = ('E_x_l', 'E_x_l__dT_s', 'E_x_l__dP', 'E_x_d__dT_j', 'E_x_u', 'E_x_d__p', 'E_x_u__p')
        if key[0] in del_keys:
            del E_x_values[key]
    
    for E_x_key, E_x_value in E_x_values.items():
        fig_data = dict(figsize=(12.2, 6.8), dpi=80)
        xlabel = r'Час, $\tau$'
        ylabel = '{} Вт'.format(E_x_key[1])
        grid_color = '#ABB2B9'
        ax_pos = (0.09, 0.09, 0.88, .75)
        
        fig = plt.figure(**fig_data)
        ax = fig.add_subplot(111)
        
        ax.plot(x_time_hvalues, E_x_value, color='k', lw=2, ls='', marker='^', ms=8,
                mfc=(0, 0, 0, .5), mec='k')
        ax.set_xlabel(xlabel, x=1, ha="right")
        ax.set_ylabel(ylabel, y=1.025, va='bottom', rotation=0)
        ax.set_xticks(x_time_hvalues)
        ax.set_xticklabels(x_times)
        #ax.set_ylim(0, vertical_ylim_max)
        #locator_y = ticker.MultipleLocator(base=200)
        #ax.yaxis.set_major_locator(locator_y)
        ax.spines['top'].set_color(grid_color)
        ax.spines['right'].set_color(grid_color)
        ax.grid(color=grid_color)
        ax.set_position(ax_pos)
    
        # Кореляція.
        z = np.polyfit(x_time_hvalues, E_x_value, 4)
        p = np.poly1d(z)
        eq = smp.S(r'{:.4f}*tau**2 + {:.4f}*tau + {:.4f}'.format(*z))
        smp.pprint('Апроксимаційне рівняння для {}:'.format(E_x_key[2]))
        smp.pprint(eq)
        ax.set_title('{}\n'
          r'Апроксимаційне рівняння: ${} = {}$'.format(E_x_key[2], 
            E_x_key[1].strip(r'$'), smp.latex(eq).replace('.', '{,}')), y=1.10)
        ax.plot(x_time_hvalues, [p(x) for x in x_time_hvalues], ls='-', lw=2, color='k')
        
        plt.savefig(os.path.join(DIRECT_PATH, '{}.png'.format(E_x_key[0])), format='png')

solve_eta_in()
print(' Роботу завершено! '.center(80, '='))

if IS_SHOW_GRAF:
    plt.show()
# -*- coding: utf-8 -*-
'''
Created on 24.02.2018

@author: ichet
'''

# Кількість сумарної (прямої та розсіяної) сонячної радіації, що надходить 
# на горизонтальну і вертикальні поверхні

import locale
locale.setlocale(locale.LC_ALL, '')
import os
from datetime import datetime
from itertools import cycle
import shutil
import numpy as np
from openpyxl import load_workbook
import sympy as smp

from matplotlib import pyplot as plt, rcParams, ticker
from matplotlib.colors import to_rgb
rcParams['axes.formatter.use_locale'] = True
rcParams.update({'font.family':'serif', 'mathtext.fontset': 'dejavuserif',
                 'font.size': 14, 'axes.titlesize' : 14})
#plt.rcParams['mathtext.fontset'] = 'dejavuserif'
#rc('font',**{'family':'serif'})
#rcParams['font.family'] = 'serif'
#rc('text', usetex=True)
#rc('text.latex',unicode=True)
#rc('text.latex',preamble=[r'\usepackage[utf8]{inputenc}\usepackage[russian]{babel}')
#rc('text.latex',preamble=r'\usepackage[russian]{babel}')

#rcParams['text.latex.preamble'] = r'\usepackage{cmap}'
#rcParams['text.latex.preamble'] = r'\usepackage{pscyr}'
#rcParams['text.latex.preamble'] = r'\usepackage{amssymb,amsfonts,amsmath,amsthm}'
#rcParams['text.latex.preamble'] = r'\usepackage[T2A]{fontenc}'
#rcParams['text.latex.preamble'] = r'\usepackage[utf8]{inputenc}'
#rcParams['text.latex.preamble'] = r'\usepackage[english,russian,ukrainian]{babel}'

#rcParams['text.latex.preamble'] = r'\usepackage{icomma}'
                                  
np.set_printoptions(precision=4)

IS_SHOW_GRAF = False
DIRECT_PATH = r'C:\Users\ichet\Dropbox\myDocs\St\Graf_Q_rad_2'
try:
    shutil.rmtree(DIRECT_PATH)
except FileNotFoundError: pass
os.mkdir(DIRECT_PATH)

print('Розрахунок проводиться в {}.'.format(datetime.strftime(datetime.now(),
                                                          '%d.%m.%Y %H:%M:%S')))    

wb = load_workbook(filename=r'C:\Users\ichet\OneDrive\Documents\DBNs\ДСТУ-Н Б В.1.1-27_2010.xlsx', 
                   read_only=True, data_only=True)
ws = wb['Лист1']

directs = np.array([c.value for c in ws['E5':'L5'][0]])
directs_iter = cycle(np.append(directs, 'Горизонтальна'))

fig_data = dict(figsize=(12.2, 6.8), dpi=80)
xlabel = 'Місяць року'
ylabel = r'$\mathrm{Q,\,\frac{кВт\cdot год}{м^2}}$'
grid_color = '#ABB2B9'
ax_pos = (0.09, 0.09, 0.88, .75)

def Q_rad(citye_name, city_r, vertical_ylim_max, full_ylim_max, reflected_ylim_max, 
          reflected_m_ylim_max, latitude, albedo):
    months = np.array([c[0].value for i, c in enumerate(ws['D{}'.format(city_r):
                                                           'D{}'.format(city_r+22)]) 
                       if not i % 2])
    month_numbers = np.arange(1, len(months)+1)
    
    city = ws['A{}'.format(city_r)].value
    print(' Сонячна радіація для м. {} '.format(city).center(80, '*'))
    
    vert_rad = np.array([[int(c.value) for c in r] for r in 
                         ws['E{}'.format(city_r):'L{}'.format(city_r+23)]]).transpose() / 3.6
    
    full_vert_rad = vert_rad[..., ::2] + vert_rad[..., 1::2]
    full_vert_rad_cumsum = np.cumsum(full_vert_rad, axis=1)
    full_vert_rad_cumsum_title = ('Кількість сумарної (прямої та розсіяної) '
        'сонячної радіації, що надходить на вертикальної поверхню '
        'за середніх умов хмарності для м. {} (широта - {} град. пн.ш., альбедо - {}%)'
            .format(city, latitude, albedo))
    print(full_vert_rad_cumsum_title + ':\n', full_vert_rad_cumsum)
    
    horizont_rad = np.array([[int(c.value) for c in r] for r in 
                             ws['M{}'.format(city_r): 'M{}'.format(city_r+23)]]) / 3.6
    horizont_rad = np.transpose(horizont_rad[::2] + horizont_rad[1::2]).ravel()
    horizont_rad_cumsum = np.cumsum(horizont_rad)
    horizont_rad_cumsum_title = ('Кількість сумарної (прямої та розсіяної) '
        'сонячної радіації, що надходить на горизонтальну поверхню '
        'за середніх умов хмарності для м. {} (широта - {} град. пн.ш., альбедо - {}%)'
            .format(city, latitude, albedo))
    print(horizont_rad_cumsum_title + ':\n', horizont_rad_cumsum)
    
    hv_rad_cumsum = np.vstack((full_vert_rad_cumsum, horizont_rad_cumsum))
    full_rad_cumsum = hv_rad_cumsum.sum(axis=0)
    full_rad_cumsum_title = ('Кількість сумарної (прямої та розсіяної) '
        'сонячної радіації, що надходить на горизонтальну і вертикальні поверхні\n'
        'за середніх умов хмарності для м. {} (широта - {} град. пн.ш., альбедо - {}%)'
            .format(city, latitude, albedo))
    print(full_rad_cumsum_title + ':\n', full_rad_cumsum)
    # Кореляція.
    z = np.polyfit(month_numbers, full_rad_cumsum, 2)
    p = np.poly1d(z)
    eq = smp.S('{:.2f}*x**2 + {:.2f}*x + {:.2f}'.format(*z))
    print('Апроксимаційне рівняння для повної радіації на горизонтальну '
          'і вертикальну поверхні:')
    smp.pprint(eq)
    
    reflected_rad_cumsum = albedo/100 * horizont_rad_cumsum
    reflected_rad_cumsum_title = ('Кількість відбитої '
        'сонячної радіації за середніх умов хмарності\nдля м. {} '
        '(широта - {} град. пн.ш., альбедо - {}%)'.format(city, latitude, albedo))
    print(reflected_rad_cumsum_title + ':\n', reflected_rad_cumsum)
    # Кореляція.
    z_2 = np.polyfit(month_numbers, reflected_rad_cumsum, 2)
    p_2 = np.poly1d(z_2)
    eq_2 = smp.S('{:.2f}*x**2 + {:.2f}*x + {:.2f}'.format(*z_2))
    print('Апроксимаційне рівняння для відбитої радіації:')
    smp.pprint(eq_2)
    
    reflected_rad_m = albedo/100 * horizont_rad
    reflected_rad_m_title = ('Кількість відбитої '
        'сонячної радіації за середніх умов хмарності\nдля м. {} '
        '(широта - {} град. пн.ш., альбедо - {}%)'.format(city, latitude, albedo))
    print(reflected_rad_m_title + ':\n', reflected_rad_m)
    # Кореляція.
    z_3 = np.polyfit(month_numbers, reflected_rad_m, 2)
    p_3 = np.poly1d(z_3)
    eq_3 = smp.S('{:.2f}*x**2 + {:.2f}*x + {:.2f}'.format(*z_3))
    print('Апроксимаційне рівняння для відбитої радіації:')
    smp.pprint(eq_3)
    
    full_and_reflect_rad_cumsum = full_rad_cumsum + reflected_rad_cumsum
    full_and_reflect_rad_cumsum_title = ('Кількість загальної (прямої, розсіяної та відбитої) '
        'сонячної радіації за середніх умов хмарності\nдля м. {} '
        '(широта - {} град. пн.ш., альбедо - {}%)'.format(city, latitude, albedo))
    print(full_and_reflect_rad_cumsum_title + ':\n', full_and_reflect_rad_cumsum)
    # Кореляція.
    z_4 = np.polyfit(month_numbers, full_and_reflect_rad_cumsum, 2)
    p_4 = np.poly1d(z_4)
    eq_4 = smp.S('{:.2f}*x**2 + {:.2f}*x + {:.2f}'.format(*z_4))
    print('Апроксимаційне рівняння для загальної радіації (комулятивної):')
    smp.pprint(eq_4)
    
    hv_rad_m = np.vstack((full_vert_rad, horizont_rad))
    full_rad_m = hv_rad_m.sum(axis=0)
    full_and_reflect_rad_m = full_rad_m + reflected_rad_m
    full_and_reflect_rad_m_title = ('Кількість загальної (прямої, розсіяної та відбитої) '
        'сонячної радіації за середніх умов хмарності\nдля м. {} '
        '(широта - {} град. пн.ш., альбедо - {}%)'.format(city, latitude, albedo))
    print(full_and_reflect_rad_m_title + ':\n', full_and_reflect_rad_m)
    # Кореляція.
    z_5 = np.polyfit(month_numbers, full_and_reflect_rad_m, 2)
    p_5 = np.poly1d(z_5)
    eq_5 = smp.S('{:.2f}*x**2 + {:.2f}*x + {:.2f}'.format(*z_5))
    print('Апроксимаційне рівняння для загальної радіації по місяцях:')
    smp.pprint(eq_5)
    
    print(u' Побудова графіків '.center(80, '='))
    fig = plt.figure(**fig_data)
    ax = fig.add_subplot(111)
    colors = cycle(('blue', '#D35400', 'green', 'red', 'purple', 'brown', 
                    'magenta', '#424949', '#117864'))
    markers = cycle(('o', '^', 'v', 's', 'P', 'x', 'd', 'p', '<'))
    for i, Qs in enumerate(hv_rad_cumsum):
        # Кореляція.
        z = np.polyfit(month_numbers, Qs, 2)
        p = np.poly1d(z)
        direct = next(directs_iter)
        eq = smp.S('{:.2f}*x**2 + {:.2f}*x + {:.2f}'.format(*z))
        #eq_g = '${:.2f}\cdot x^2 + {:.2f}\cdot x + {:.2f}$'.format(*z)
        print('Апроксимаційне рівняння для напрямку {}:'.format(direct))
        smp.pprint(eq)
        
        color = next(colors)
        marker = next(markers)
        ax.plot(month_numbers, Qs, color=color, marker=marker, ms=8,
                mfc=to_rgb(color)+(.5,), mec=color)
        if i in (0, 6, 7, 8):
            k = -4
            tx = 1
            ty = -28
        elif i in (1, 4):
            k = -5
            tx = -1
            ty = 50
        elif i in (2,):
            k = -2
            tx = .8
            ty = 83
        elif i in (3,):
            k = -2
            tx=.8
            ty = 70
        elif i in (5,):
            k = -4
            tx = 1
            ty = -14
            
        annot_x = (month_numbers[k] + month_numbers[k+1]) / 2.
        annot_y = (Qs[k]+Qs[k+1]) / 2.
        ax.annotate(direct, xy=(annot_x, annot_y), xycoords='data',
                    xytext=(annot_x+tx, annot_y+ty), textcoords='data', va='top', 
                    ha='left', arrowprops=dict(arrowstyle='-'))
    
        #print()
        #ax.plot(month_numbers, [p(x) for x in month_numbers], ls='--', lw=0.8,
        #        color=color, marker=marker)
    
    def save_graf(city_name, view_name):
        plt.savefig(os.path.join(DIRECT_PATH, '{}-{}-{}-{}.png'.format(
            city_name, view_name, latitude, albedo)), format='png')
    
    ax.set_title('Кількість сумарної (прямої та розсіяної) '
        'сонячної радіації, що надходить на поверхню\n'
        'за середніх умов хмарності для м. {} (широта - {} град. пн.ш., альбедо - {}%)'
                  .format(city, latitude, albedo), y=1.11)
    ax.set_xlabel(xlabel, x=1, ha="right")
    ax.set_ylabel(ylabel, y=1.025, va='bottom', rotation=0)
    ax.set_xticks(month_numbers)
    ax.set_ylim(0, vertical_ylim_max)
    locator_y = ticker.MultipleLocator(base=200)
    ax.yaxis.set_major_locator(locator_y)
    ax.spines['top'].set_color(grid_color)
    ax.spines['right'].set_color(grid_color)
    ax.grid(color=grid_color)
    ax.set_position(ax_pos)
    save_graf(citye_name, 'compos')
    
    # На вертикальну і горизонтальну поверхню.
    fig = plt.figure(**fig_data)
    ax = fig.add_subplot(111)
    ax.plot(month_numbers, full_rad_cumsum, color='k', marker='^', ms=8,
            mfc=(0, 0, 0, .5), mec='k')
    ax.set_title(full_rad_cumsum_title, y=1.11)
    ax.set_xlabel(xlabel, x=1, ha='right')
    ax.set_ylabel(ylabel, y=1.025, va='bottom', rotation=0)
    ax.set_xticks(month_numbers)
    ax.set_ylim(0, full_ylim_max)
    locator_y = ticker.MultipleLocator(base=500)
    ax.yaxis.set_major_locator(locator_y)
    ax.spines['top'].set_color(grid_color)
    ax.spines['right'].set_color(grid_color)
    ax.grid(color=grid_color)
    ax.set_position(ax_pos)
    save_graf(citye_name, 'full')
    
    # На відбиту поверхню.
    fig = plt.figure(**fig_data)
    ax = fig.add_subplot(111)
    ax.plot(month_numbers, reflected_rad_cumsum, color='k', marker='^', ms=8,
            mfc=(0, 0, 0, .5), mec='k')
    ax.set_title(reflected_rad_cumsum_title, y=1.11)
    ax.set_xlabel(xlabel, x=1, ha='right')
    ax.set_ylabel(ylabel, y=1.025, va='bottom', rotation=0)
    ax.set_xticks(month_numbers)
    ax.set_ylim(0, reflected_ylim_max)
    locator_y = ticker.MultipleLocator(base=100)
    ax.yaxis.set_major_locator(locator_y)
    ax.spines['top'].set_color(grid_color)
    ax.spines['right'].set_color(grid_color)
    ax.grid(color=grid_color)
    ax.set_position(ax_pos)
    save_graf(citye_name, 'reflect')
    
    # На відбиту поверхню по місяцях.
    fig = plt.figure(**fig_data)
    ax = fig.add_subplot(111)
    ax.bar(month_numbers, reflected_rad_m)
    ax.plot(month_numbers, reflected_rad_m, color='r', ls='--', marker='^', ms=8,
            mfc=(0, 0, 0, .5), mec='k')
    ax.set_title(reflected_rad_m_title, y=1.11)
    ax.set_xlabel(xlabel, x=1, ha='right')
    ax.set_ylabel(ylabel, y=1.025, va='bottom', rotation=0)
    ax.set_xticks(month_numbers)
    ax.set_ylim(0, reflected_m_ylim_max)
    locator_y = ticker.MultipleLocator(base=10)
    ax.yaxis.set_major_locator(locator_y)
    ax.spines['top'].set_color(grid_color)
    ax.spines['right'].set_color(grid_color)
    ax.grid(color=grid_color)
    ax.set_position(ax_pos)
    save_graf(citye_name, 'reflect_m')
    
    # Сумарна (пряма+розсіяна+відбита) комулятивна.
    fig = plt.figure(**fig_data)
    ax = fig.add_subplot(111)
    ax.plot(month_numbers, full_and_reflect_rad_cumsum, color='k', marker='^', ms=8,
            mfc=(0, 0, 0, .5), mec='k')
    ax.set_title(full_and_reflect_rad_cumsum_title, y=1.11)
    ax.set_xlabel(xlabel, x=1, ha='right')
    ax.set_ylabel(ylabel, y=1.025, va='bottom', rotation=0)
    ax.set_xticks(month_numbers)
    locator_y_base = 500
    ax.set_ylim(0, get_y_max(full_and_reflect_rad_cumsum, locator_y_base))
    locator_y = ticker.MultipleLocator(base=locator_y_base)
    ax.yaxis.set_major_locator(locator_y)
    ax.spines['top'].set_color(grid_color)
    ax.spines['right'].set_color(grid_color)
    ax.grid(color=grid_color)
    ax.set_position(ax_pos)
    save_graf(citye_name, 'summary')
    
    # Сумарна (пряма+розсіяна+відбита) по місяцях.
    fig = plt.figure(**fig_data)
    ax = fig.add_subplot(111)
    ax.bar(month_numbers, full_and_reflect_rad_m)
    ax.plot(month_numbers, full_and_reflect_rad_m, color='r', ls='--', 
            marker='^', ms=8, mfc=(0, 0, 0, .5), mec='k')
    ax.set_title(full_and_reflect_rad_m_title, y=1.11)
    ax.set_xlabel(xlabel, x=1, ha='right')
    ax.set_ylabel(ylabel, y=1.025, va='bottom', rotation=0)
    ax.set_xticks(month_numbers)
    locator_y_base = 100
    ax.set_ylim(0, get_y_max(full_and_reflect_rad_m, locator_y_base))
    locator_y = ticker.MultipleLocator(base=locator_y_base)
    ax.yaxis.set_major_locator(locator_y)
    ax.spines['top'].set_color(grid_color)
    ax.spines['right'].set_color(grid_color)
    ax.grid(color=grid_color)
    ax.set_position(ax_pos)
    save_graf(citye_name, 'summary_m')
    
    return full_and_reflect_rad_cumsum, full_and_reflect_rad_m

def get_y_max(y, y_locator_base):
    y_min = np.floor(min(y))
    y_min_delta = y_min % y_locator_base
    if y_min_delta:
        y_min -= y_min_delta
    y_max = np.ceil(max(y))
    y_max_k, y_max_delta = divmod(y_max, y_locator_base)
    if y_max_delta:
        y_max = y_locator_base * (y_max_k+1)
    
    return y_max

month_numbers = np.arange(1, 13)

# Сумарні (прямі+розсіяні+відбиті) комулятивні на одному графіку.
fig_cumsum = plt.figure(**fig_data)
ax_cumsum = fig_cumsum.add_subplot(111)
ax_cumsum.set_title('Кількість загальної (прямої, розсіяної та відбитої) '
             'сонячної радіації за середніх умов хмарності', y=-0.19)
ax_cumsum.set_xlabel(xlabel, x=1, ha='right')
ax_cumsum.set_ylabel(ylabel, y=1.025, va='bottom', rotation=0)
ax_cumsum.set_xticks(month_numbers)
locator_y_cumsum_base = 500
locator_y = ticker.MultipleLocator(base=locator_y_cumsum_base)
ax_cumsum.yaxis.set_major_locator(locator_y)
ax_cumsum.spines['top'].set_color(grid_color)
ax_cumsum.spines['right'].set_color(grid_color)
ax_cumsum.grid(color=grid_color)
ax_cumsum.set_position((0.09, .14, .88, .72)) # (0.09, 0.09, 0.88, .75)

# Сумарні (прямі+розсіяні+відбиті) по місяцях на одному графіку.
fig_m = plt.figure(**fig_data)
ax_m = fig_m.add_subplot(111)
ax_m.set_title('Кількість загальної (прямої, розсіяної та відбитої) '
               'сонячної радіації за середніх умов хмарності по місяцях', y=-0.19)
ax_m.set_xlabel(xlabel, x=1, ha='right')
ax_m.set_ylabel(ylabel, y=1.025, va='bottom', rotation=0)
ax_m.set_xticks(month_numbers)
locator_y_m_base = 100
locator_y = ticker.MultipleLocator(base=locator_y_m_base)
ax_m.yaxis.set_major_locator(locator_y)
ax_m.spines['top'].set_color(grid_color)
ax_m.spines['right'].set_color(grid_color)
ax_m.grid(color=grid_color)
ax_m.set_position((0.09, .14, .88, .72))

Q_cumsum = np.array([])
lines_cumsum = []
Q_m = np.array([])
lines_m = []
labels = []
for (Q_cumsum_i, Q_m_i), linestyle, color, marker, label in zip(
                      (Q_rad('Kiev', 150, 1200, 5500, 500, 72, 50.24, 38.5), # Для Києва 50,24.
                      Q_rad('Lviv', 246, 1000, 5000, 400, 62, 49.49, 38), # Для Львова 49,49.
                      Q_rad('Harkiv', 462, 1200, 5550, 400, 62, 49.56, 35), # Для Харкова 49,56.
                      Q_rad('Simpheropol', 366, 1400, 6000, 400, 62, 44.57, 31), # Для Сімферополь 44,57.
                      Q_rad('Odessa', 294, 1400, 6000, 500, 72, 46.26, 33), # Для Одеса 46,26.
                      Q_rad('Zaporizza', 102, 1200, 6000, 400, 62, 47.48, 32.5), # Для Запоріжжя 47,48.
                      Q_rad('Chernivci', 558, 1050, 5000, 400, 62, 48.16, 36), # Для Чернівці - 48,16.
                      Q_rad('Chernigiv', 582, 1200, 5500, 500, 72, 51.24, 40)),
                      ('-', '--', '-.', ':', (0, (3, 1, 1, 1, 1, 1)),
                       (0, (3, 5, 1, 5, 1, 5)),  (0, (3, 10, 1, 10, 1, 10)),
                       (0, (3, 10, 1, 10))),
                      ('blue', '#D35400', 'green', 'red', 'purple', 'brown', 
                       'magenta', '#424949', '#117864'),
                      ('o', '^', 'v', 's', 'P', 'x', 'd', 'p', '<'),
                      ('Київ', 'Львів', 'Харків', 'Сімферополь', 'Одеса',
                       'Запоріжжя', 'Чернівці', 'Чернігів')): # Для Чернігів - 51,24.
    line, = ax_cumsum.plot(month_numbers, Q_cumsum_i, ls=linestyle, color=color,
                           marker=marker, ms=8, mfc=(0, 0, 0, .5), mec='k')
    lines_cumsum.append(line)
    Q_cumsum = np.r_[Q_cumsum, Q_cumsum_i]
    
    line, = ax_m.plot(month_numbers, Q_m_i, ls=linestyle, color=color, 
                      marker=marker, ms=8, mfc=(0, 0, 0, .5), mec='k')
    lines_m.append(line)
    Q_m = np.r_[Q_m, Q_m_i]
    
    labels.append(label)

ax_cumsum.set_ylim(0, get_y_max(Q_cumsum, locator_y_cumsum_base))
ax_cumsum.legend(loc=3, ncol=8, mode='expand', bbox_to_anchor=(.03, 1.10, .92, .102),
                  handles=lines_cumsum, labels=labels, fontsize=12)
fig_cumsum.savefig(os.path.join(DIRECT_PATH, 'summary_one_cumsum.png'), format='png')

ax_m.set_ylim(0, get_y_max(Q_m, locator_y_m_base))
ax_m.legend(loc=3, ncol=8, mode='expand', bbox_to_anchor=(.03, 1.10, .92, .102),
                  handles=lines_m, labels=labels, fontsize=11)
fig_m.savefig(os.path.join(DIRECT_PATH, 'summary_one_m.png'), format='png')

print(' Роботу завершено! '.center(80, '='))

if IS_SHOW_GRAF:
    plt.show()
# -*- coding: utf-8 -*-

'''
Created on 26.02.2018

@author: ichet
'''

import locale
locale.setlocale(locale.LC_ALL, '') # Для сортування.
from copy import deepcopy
import time

import numpy as np
from openpyxl import load_workbook

wb = load_workbook(filename=r'C:\Users\ichet\Dropbox\myDocs\MySite\weather_in_ua.xlsx', 
                   read_only=True, data_only=True)
ws = wb.get_sheet_by_name('Weather_un_ua_UKR')

print(' Викачані дані: '.center(80, '='))
data = np.array([[c.value for c in r] for r in ws['B2':'H28689']])
print('Всього елементів:', len(data))

create_dict_1 = lambda data, key=locale.strxfrm: {v:i for i, v in enumerate(
                                                   sorted(data, key=key), start=1)}
create_dict_2 = lambda data, key=locale.strxfrm: {tuple(v):i for i, v in enumerate(
                                                   sorted(data, key=key), start=1)}

print(' Статуси: '.center(80, '='))
statuses = np.sort(np.unique(data[..., 1])) # Унікальні статуси
print('Всього - {}; {}'.format(len(statuses), statuses))
status_alias = {'-': 'NULL', 'місто': 'м.', 'село': 'с.', 'селище': 'с.', 
                'селище міського типу': 'смт', 'смт': 'смт'}

print(' Області: '.center(80, '='))
regions = create_dict_1(np.unique(data[..., 2]))    # Унікальні області.
print('Всього - {}; {}'.format(len(regions), regions))

print(' Райони: '.center(80, '='))
areas = create_dict_2(np.unique(data[..., (2, 3)], axis=0).tolist(), 
                    lambda v: locale.strxfrm(v[1])) # Унікальні пари - область і район.
print('Всього - {}; {}'.format(len(areas), areas))

village_and_geodata = np.array([((i, areas[r, a], v, status_alias[s]),
                     (i, float(lat), float(lon), int(alt))) 
                     for i, (v, s, r, a, lat, lon, alt) in enumerate(data, start=1)],
                     dtype=np.object)

print(' Для перейменування географічних пунтків: '.center(80, '*'))
#regions_copy = deepcopy(regions)
areas_copy = deepcopy(areas)    # Глибока копія для районів.
villages_copy = deepcopy(village_and_geodata[:, 0, :])  # Глибока копія для міст.
wb2 = load_workbook(r'C:\Users\ichet\Dropbox\myDocs\MySite\Декомунізація_27-02-2018.xlsx',
                    read_only=True, data_only=True)
ws2_area = wb2.get_sheet_by_name('Райони')
print(' Викачані дані районів для перейменування: '.center(80, '='))
data_rename = np.array([[c.value for c in r] for r in ws2_area['A1':'C20']])
print(len(data_rename))
print(' Змінені райони: '.center(80, '='))
old_areas = {}
for oldarea, region, newarea in data_rename:
    key = (region, oldarea)
    value = areas_copy.pop(key)
    old_areas[key] = value
    areas_copy[region, newarea] = value
print('Всього нових: {}; {}'.format(len(areas_copy), areas_copy))
print('Всього старих: {}; {}'.format(len(old_areas), old_areas))

ws2_village = wb2.get_sheet_by_name('Населені пункти')
data_rename = {}
for row in ws2_village['A1':'D1011']:
    if not row[1].value:
        rk = row[0].value   # Назва області.
        data_rename[rk] = []
    else:
        data_rename[rk].append([c.value for c in row])
print(' Змінені географічні пункти: '.center(80, '='))
old_villages = set()
# Для рядків таблиці.
villages_copy_dict = {tuple(v[1:]): i for i, v in enumerate(villages_copy)}
for region, data in data_rename.items():
    for status, oldvillage, area, newvillage in data:
        status = status_alias[status]
        id_area = areas[region, area]
        # Назва області і району, назва міста і статус - унікальні.
        # Але бувають однакові населені пункти в одній області, районі.
        r_village = villages_copy_dict[(id_area, oldvillage, status)]
        old_status = villages_copy[r_village, 3]
        villages_copy[r_village, (2, 3)] = newvillage, status
        old_villages.add((villages_copy[r_village, 0], oldvillage, old_status))

print(' Створення дампу бази даних: '.center(80, '='))
str_st = ','.join('"{}"'.format(s) for s in 
                  set(status_alias.values()) if s != 'NULL')

ua_region = '''/* Дамп міст та сіл створено програмно мовою Python від %s */

DROP TABLE IF EXISTS `ua_region`;
CREATE TABLE `ua_region` (
    `id_region` TINYINT(2) UNSIGNED NOT NULL PRIMARY KEY AUTO_INCREMENT,
    `region` VARCHAR(60) UNIQUE NOT NULL
) ENGINE=InnoDB DEFAULT CHARSET=utf8;

LOCK TABLES `ua_region` WRITE;
/*!40000 ALTER TABLE `ua_region` DISABLE KEYS */;
INSERT INTO `ua_region` (`id_region`, `region`) VALUES
%s;
/*!40000 ALTER TABLE `ua_region` ENABLE KEYS */;
UNLOCK TABLES;''' % (time.ctime(time.time()),
                     ',\n'.join(str((k, v)) for v, k in regions.items()))

ua_area = '''
DROP TABLE IF EXISTS `ua_area`;
CREATE TABLE `ua_area` (
    `id_area`  SMALLINT(3) UNSIGNED NOT NULL PRIMARY KEY AUTO_INCREMENT,
    `id_region` TINYINT(2) UNSIGNED NOT NULL,
    `area` VARCHAR(60) NOT NULL,
    UNIQUE KEY `ufk_area` (`id_area`, `id_region`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8;

LOCK TABLES `ua_area` WRITE;
/*!40000 ALTER TABLE `ua_area` DISABLE KEYS */;
INSERT INTO `ua_area` (`id_area`, `id_region`, `area`) VALUES
%s;
/*!40000 ALTER TABLE `ua_area` ENABLE KEYS */;
UNLOCK TABLES;''' % ',\n'.join(str((k, regions[v[0]], v[1])) for v, k in 
                               areas_copy.items())

ua_village = '''
DROP TABLE IF EXISTS `ua_village`;
CREATE TABLE `ua_village` (
    `id_village` SMALLINT UNSIGNED NOT NULL PRIMARY KEY AUTO_INCREMENT,
    `id_area` SMALLINT(3) UNSIGNED NOT NULL,
    `village` VARCHAR(60) NOT NULL,
    `status` ENUM(%s),
    UNIQUE KEY `ufk_village` (`id_village`, `id_area`, `status`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8;

LOCK TABLES `ua_village` WRITE;
/*!40000 ALTER TABLE `ua_village` DISABLE KEYS */;
INSERT INTO `ua_village` (`id_village`, `id_area`,`village`, `status`) VALUES
%s;
/*!40000 ALTER TABLE `ua_village` ENABLE KEYS */;
UNLOCK TABLES;''' % (str_st, ',\n'.join(str(tuple(v)) for v in villages_copy),)

ua_geocoords = '''
DROP TABLE IF EXISTS `ua_geocoords`;
CREATE TABLE `ua_geocoords` (
    `id_village` SMALLINT UNSIGNED NOT NULL PRIMARY KEY,
    `latitude` FLOAT(6,2) NOT NULL,
    `longitude` FLOAT(7,2) NOT NULL,
    `altitude` SMALLINT NOT NULL
) ENGINE=InnoDB DEFAULT CHARSET=utf8;

LOCK TABLES `ua_geocoords` WRITE;
/*!40000 ALTER TABLE `ua_geocoords` DISABLE KEYS */;
INSERT INTO `ua_geocoords` (`id_village`, `latitude`, `longitude`, `altitude`) VALUES
%s;
/*!40000 ALTER TABLE `ua_geocoords` ENABLE KEYS */;
UNLOCK TABLES;''' % ',\n'.join(str(tuple(v)) for v in village_and_geodata[:, 1])

ua_oldregion = '''
DROP TABLE IF EXISTS `ua_oldregion`;
CREATE TABLE `ua_oldregion` (
    `id_region` TINYINT(2) UNSIGNED NOT NULL,
    `oldregion` VARCHAR(60) NOT NULL,
    PRIMARY KEY (`id_region`, `oldregion`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8;'''

#'''
#LOCK TABLES `ua_oldregion` WRITE;
#/*!40000 ALTER TABLE `ua_oldregion` DISABLE KEYS */;
#INSERT INTO `ua_oldregion` (`id_region`, `oldregion`) VALUES
#%s;
#/*!40000 ALTER TABLE `ua_oldregion` ENABLE KEYS */;
#UNLOCK TABLES;''' % ''

ua_oldarea = '''
DROP TABLE IF EXISTS `ua_oldarea`;
CREATE TABLE `ua_oldarea` (
    `id_area`  SMALLINT(3) UNSIGNED NOT NULL,
    `oldarea` VARCHAR(60) NOT NULL,
    PRIMARY KEY (`id_area`, `oldarea`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8;

LOCK TABLES `ua_oldarea` WRITE;
/*!40000 ALTER TABLE `ua_oldarea` DISABLE KEYS */;
INSERT INTO `ua_oldarea` (`id_area`, `oldarea`) VALUES
%s;
/*!40000 ALTER TABLE `ua_oldarea` ENABLE KEYS */;
UNLOCK TABLES;''' % ',\n'.join(str((k, v[1])) for v, k in old_areas.items())

ua_oldvillage = '''
DROP TABLE IF EXISTS `ua_oldvillage`;
CREATE TABLE `ua_oldvillage` (
    `id_village` SMALLINT UNSIGNED NOT NULL,
    `oldvillage` VARCHAR(60) NOT NULL,
    `status` ENUM(%s),
    PRIMARY KEY (`id_village`, `oldvillage`, `status`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8;

LOCK TABLES `ua_oldvillage` WRITE;
/*!40000 ALTER TABLE `ua_oldvillage` DISABLE KEYS */;
INSERT INTO `ua_oldvillage` (`id_village`, `oldvillage`, `status`) VALUES
%s;
/*!40000 ALTER TABLE `ua_oldvillage` ENABLE KEYS */;
UNLOCK TABLES;''' % (str_st, ',\n'.join(str(tuple(v)) for v in old_villages))

dump_region = ''.join((ua_region, ua_area, ua_village, ua_geocoords, ua_oldregion,
                       ua_oldarea, ua_oldvillage))

# Заміна відустніх значень.
dump_region = dump_region.replace("'NULL'", 'NULL')
dump_region = dump_region.replace("(8651, 1, 'Новопетрівське', 'с.')",
                                  "(8651, 1, 'Садове', 'с.')")
dump_region = dump_region.replace("(8652, 'Новопетрівське', 'с.')",
                                  "(8651, 'Новопетрівське', 'с.')", 1)

dump_region = dump_region.replace("(2395, 7, 'Красний Пахар', 'с.')",
                                  "(2395, 7, 'Ступакове', 'с.')")
dump_region = dump_region.replace("(2396, 'Красний Пахар', 'с.')",
                                  "(2395, 'Красний Пахар', 'с.')", 1)

fname = 'dump_location.sql'
with open(r'C:\Users\ichet\Dropbox\myDocs\MySite\\' + fname, mode='w', 
          encoding='utf-8') as f:
    f.write(dump_region)
print(' Створення дампу "{}" завершено! '.format(fname).center(80, '*'))
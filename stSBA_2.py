# -*- coding: utf-8 -*-

'''
Created on 8 трав. 2018 р.

@author: ichet
'''

import locale
locale.setlocale(locale.LC_ALL, '')
import os
import numpy as np
import shutil
from enum import Enum
from datetime import datetime
from openpyxl import load_workbook
from scipy.interpolate import interp1d
import sympy as smp
from matplotlib import pyplot as plt, rcParams, ticker

rcParams['axes.formatter.use_locale'] = True
rcParams.update({'font.family':'serif', 'mathtext.fontset': 'dejavuserif',
                 'font.size': 16, 'axes.titlesize' : 16})

class TargetGraph(Enum):
    no_showsave = 0
    save = 1
    show = 2
    show_and_save = 3

target_graph = TargetGraph.show_and_save
if target_graph in (TargetGraph.save, TargetGraph.show_and_save):
    DIRECT_PATH = r'C:\Users\ichet\Dropbox\myDocs\St\I_maxs'
    try:
        # Спочатку видаляється папка.
        shutil.rmtree(DIRECT_PATH)
    except FileNotFoundError: pass
    os.mkdir(DIRECT_PATH)   # Потім створюється папка.

fig_data = dict(figsize=(12.5, 6), dpi=100)
grid_color = '#ABB2B9'
ax_pos = (0.10, .16, .89, .755)

m = np.arange(1, 13)

wb = load_workbook(filename=r'C:\Users\ichet\OneDrive\Documents\DBNs\ДСТУ-Н Б В.1.1-27_2010.xlsx', 
                   read_only=True, data_only=True)
ws = wb.get_sheet_by_name('Т10-Т17.Енергетична осв. пов.')

ws_2 = wb.get_sheet_by_name('Лист2')
cities = ('Запоріжжя', 'Київ', 'Львів', 'Одеса', 'Сімферополь', 'Харків',
          'Чернівці', 'Чернігів')
city_phi = {r[0].value: round(int(r[1].value) + int(r[2].value)/60., 2) 
                          for i, r in enumerate(ws_2['A6':'C605']) if (not i % 24) and 
                            (r[0].value in cities)}

phi_range_1 = {44: (6, 23), 46: (40, 23),  48: (74, 23), 50: (108, 19)}
phi_range_7 = {44: (138, 35), 46: (184, 35), 48: (230, 35), 50: (276, 35)}

# Тривалість світлового дня з сайту - http://ua.365.wiki/world/ukraine/lviv/sun/
city_tau_c = {'Запоріжжя': ('08:55    10:16    11:57    13:41    15:11    '
                            '15:57    15:33    14:13    12:32    10:49    '
                            '09:16    08:28'),
              'Київ': ('08:36    10:06    11:56    13:51    15:30    16:22    '
                       '15:55    14:26    12:35    10:41    08:59    08:05'),
              'Львів': ('08:41    10:08    11:56    13:48    15:25    16:16    '
                        '15:49    14:22    12:34    10:43    09:03    08:11'),
              'Одеса': ('09:05    10:21    11:57    13:36    15:02    15:45    '
                        '15:22    14:07    12:31    10:52    09:25    08:39'),
              'Сімферополь': ('09:14    10:27    11:57    13:32    14:52    '
                              '15:33    15:11    14:00    12:30    10:56    '
                              '09:33    08:50'),
              'Харків': ('08:39    10:08    11:56    13:49    15:27    16:17    '
                         '15:51    14:23    12:35    10:43    09:02    08:09'),
              'Чернівці': ('08:52    10:15    11:56    13:43    15:14    '
                           '16:01    15:36    14:15    12:33    10:47    '
                           '09:14    08:24'),
              'Чернігів': ('08:27    10:01    11:56    13:55    15:39    '
                           '16:33    16:04    14:31    12:36    10:38    '
                           '08:52    07:55')
              }

n = smp.Symbol('n')
s_E = []
Ss_E = []

def E_max(begin_r, dr):
    '''Визначає максимальну інтенсивність сонячної радіації для заданої широти
    
    @param: begin_r (str) - початковий рядок захоплення;
    @param: dr (str) - максимальна різниця зміщення;
    '''
    
    E_sm = np.array([int(r[0].value) for r in ws['J{}'.format(begin_r): 
                                                 'J{}'.format(begin_r+dr)]])
    return np.max(E_sm[::2] + E_sm[1::2])

def generate_Imax_graph(city):
    phi = city_phi[city]
    phi_list = np.array(sorted(list(phi_range_1.keys())))
    args_min = np.argwhere(phi_list <= phi)
    phi_min = phi_list[np.max(args_min)]
    
    args_max = np.argwhere(phi_list >= phi)
    if len(args_max):
        phi_max = phi_list[np.min(args_max)]
    else:
        phi_max = phi_min
        phi_min = phi_list[np.max(args_min[:-1])]
    
    if phi_min == phi_max:
        E_m_1 = E_max(*phi_range_1[phi_min])
        E_m_7 = E_max(*phi_range_7[phi_min])
    else:
        E_1_min = E_max(*phi_range_1[phi_min])
        E_1_max = E_max(*phi_range_1[phi_max])
        E_m_1 = E_1_min + (E_1_max - E_1_min) / (phi_max - phi_min) * (phi - phi_min)
        
        E_7_min = E_max(*phi_range_7[phi_min])
        E_7_max = E_max(*phi_range_7[phi_max])
        E_m_7 = E_7_min + (E_7_max - E_7_min) / (phi_max - phi_min) * (phi - phi_min)
    
    #E_m_1 = 280
    #E_m_7 = 862
    
    # Тривалість світлового дня з сайту - http://ua.365.wiki/world/ukraine/lviv/sun/
    hm = lambda h, m: round(int(h)+int(m)/60., 2)
    tau_c = {i: hm(*v.split(':')) for i, v in enumerate(city_tau_c[city].split(), start=1)}
    
    # Підігнане.
    #tau_c = {1: 8.86, 2: 10.13, 3: 11.93, 4: 13.72, 5: 15.42, 6: 16.27, 7: 15.81, 
    #                     8: 14.37, 9: 12.57, 10: 10.72, 11: 9.05, 12: 8.18}
             
    
    E_1m = {i: int(round(E_m_1 + (E_m_7 - E_m_1) / (tau_c[7] - tau_c[1]) * (tau - tau_c[1]), 0)) 
           for i, tau in tau_c.items()}
    print('Варіант 1: I_max_1 = {} Вт/м^2'.format(E_1m))
    
    E_2m = {tau: int(round(E_m_1 + (E_m_7 - E_m_1) / (np.sin(np.pi/2) - np.sin(np.pi)) * 
                 (np.sin( (np.pi/2/(7-1)*np.abs(tau-1)) if tau < 8 else  (np.pi/2 - 
                 (np.pi-np.pi/2)/(13-7)*(tau-7))) - np.sin(np.pi)), 0)) for tau in m}
    print('Варіант 2: I_max_2 = {} Вт/м^2'.format(E_2m))
    s_E.append((max(E_1m.values()), max(E_2m.values())))
    
    fig = plt.figure(**fig_data)
    ax = fig.add_subplot(111)
    
    x = m
    y_1 = np.array(list(E_1m.values()))
    
    x_a = np.linspace(x.min(), x.max(), 100)
    f = interp1d(x, y_1, kind='cubic')
    fy_1 = f(x_a)
    plt.plot(x_a, fy_1, lw=2.5, c='k')#, label=r'$\mathrm{E_m}(\tau_c)$')
    plt.plot(x, y_1, 'o', lw=2.5, c='k')
    ax.annotate(r'1', xy=(x_a[-25], f(x_a[-25])), xycoords='data',
                                xytext=(x_a[-21], f(x_a[-29])), textcoords='data', va='top', 
                                ha='left', arrowprops=dict(arrowstyle='-'))
    
    # Кореляція.
    z = np.polyfit(x, y_1, 3)
    p = np.poly1d(z)
    eq_1 = smp.S(r'{:.4f}*n**3 + {:.4f}*n**2 + {:.4f}*n + {:.4f}'.format(*z))
    
    y_2 = np.array(list(E_2m.values()))
    f = interp1d(x, y_2, kind='cubic')
    fy_2 = f(x_a)
    plt.plot(x_a, fy_2, ls='--', lw=2.5, c='k')#, label=r'$\mathrm{E_m}(\sin(\tau))$')
    plt.plot(x, y_2, 'o', marker='^', lw=2.5, c='k')
    ax.annotate(r'2', xy=(x_a[-26], f(x_a[-26])), xycoords='data',
                                xytext=(x_a[-21], f(x_a[-32])), textcoords='data', va='top', 
                                ha='left', arrowprops=dict(arrowstyle='-'))
    
    # Кореляція.
    z = np.polyfit(x, y_2, 3)
    p = np.poly1d(z)
    eq_2 = smp.S(r'{:.4f}*n**3 + {:.4f}*n**2 + {:.4f}*n + {:.4f}'.format(*z))
    
    S_E_1 = smp.integrate(eq_1, (n, 1, 12))
    S_E_2 = smp.integrate(eq_2, (n, 1, 12))
    Ss_E.append((S_E_1, S_E_2))
    delta_E = (S_E_2 - S_E_1) / S_E_2 * 100
    
    y = np.r_[fy_1, fy_2]
    
    x_locator_base = 1
    delta_x = (x[1] - x[0]) / 2
    x_min = x[0] - delta_x
    x_max = x[-1] + delta_x
    
    y_locator_base = 100
    y_min = np.floor(min(y))
    y_min_delta = y_min % y_locator_base
    if y_min_delta:
        y_min -= y_min_delta
    y_max = np.ceil(max(y))
    y_max_k, y_max_delta = divmod(y_max, y_locator_base)
    if y_max_delta:
        y_max = y_locator_base * (y_max_k+1)
    
    ax.set_xlabel('Місяці, n', x=1, ha="right")
    locator_x = ticker.MultipleLocator(base=x_locator_base)
    ax.xaxis.set_major_locator(locator_x)
    #plt.xticks(x, ['XII', 'I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VII', 'IX', 'X', 
    #                    'XI', 'XII'])
    ax.set_xlim(x_min, x_max)
    
    ax.set_ylabel(r'$\mathrm{I_{max},\ \frac{Вт}{м^2}}$', y=1.025, va='bottom', 
                  rotation=0)
    locator_y = ticker.MultipleLocator(base=y_locator_base)
    ax.yaxis.set_major_locator(locator_y)
    ax.set_ylim(y_min, y_max)
    
    title_eq_1 = smp.latex(eq_1).replace('.', '{,}')
    title_eq_2 = smp.latex(eq_2).replace('.', '{,}')
    eq_sub_title = (r'$\mathrm{I_{max}(\tau_c) = %s}$,%s$\mathrm{I_{max}(\sin(m)) = %s}$' 
                    % (title_eq_1, '\t', title_eq_2))
    
    ax.set_title(r'Зміна максимальної інтенсивності по місяцях, $\mathrm{I_{max}}$, '
                 r'$\mathrm{\frac{Вт}{м^2}}$' + r' для м. {} ($\varphi = {}^\circ$)'
                 .format(city, '{}'.format(phi).replace('.', '{,}')) + '\n' + 
                 eq_sub_title + '\t' + r'$\delta_{{E}} = {:.1f}$%'
                 .format(delta_E).replace('.', '{,}'), 
                 y=-0.215, fontsize=8)
    ax.spines['top'].set_color(grid_color)
    ax.spines['right'].set_color(grid_color)
    ax.grid(color=grid_color)
    ax.set_position(ax_pos)
    
    fig.canvas.set_window_title('I_max для м. {}'.format(city))
    
    if target_graph in (TargetGraph.save, TargetGraph.show_and_save):
        plt.savefig(os.path.join(DIRECT_PATH, 'I_max for {}.png'.format(city)), format='png', 
                    dpi=fig_data['dpi'])
    #plt.legend()

for city in cities:
    print(' Для міста {}: '.format(city).center(80, '='))
    generate_Imax_graph(city)

sE_1 = [item[0] for item in s_E]
sE_1_min = min(sE_1); sE_1_max = max(sE_1); Delta_E_1 = sE_1_max - sE_1_min
sE_2 = [item[1] for item in s_E]
sE_2_min = min(sE_2); sE_2_max = max(sE_2); Delta_E_2 = sE_2_max - sE_2_min
print("Похибка для максимальних I_max для міст: Delta_I_max_1 = {0} Вт/м^2, "
      "delta_I_max_1 = {1:.1f}%, Delta_I_max_2 = {2} Вт/м^2, delta_I_max_2 = {3:.1f}%"
      .format(Delta_E_1, Delta_E_1 / sE_1_min * 100, 
              Delta_E_2, Delta_E_2 / sE_2_min * 100) )
print('Середнє значення I_max_сер_1 = {:.0f} Вт/м^2, I_max_сер_2 = {:.0f} Вт/м^2'
      .format(sum(sE_1)/len(sE_1), sum(sE_2)/len(sE_2)))

Ss_E_1 = [item[0] for item in Ss_E]
Ss_E_1_min = min(Ss_E_1); Ss_E_1_max = max(Ss_E_1)

Ss_E_2 = [item[1] for item in Ss_E]
Ss_E_2_min = min(Ss_E_2); Ss_E_2_max = max(Ss_E_2)

print("Похибка максимальної та мінімальної енергії, E, протягом року: "
      "delta_E_1 = {:.1f}%, delta_E_2 = {:.1f}%".format(
          (Ss_E_1_max - Ss_E_1_min) / Ss_E_1_min * 100,
          (Ss_E_2_max - Ss_E_2_min) / Ss_E_2_min * 100
      ))

print(" Розв'язок завершено! ".center(50, '='))

if target_graph in (TargetGraph.show, TargetGraph.show_and_save):
    plt.show()
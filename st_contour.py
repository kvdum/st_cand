# -*- coding: utf-8 -*-

'''
Created on 23.04.2018.

@author: ichet
'''

import locale
locale.setlocale(locale.LC_ALL, '')

import os
from copy import deepcopy
from collections import namedtuple
from datetime import datetime
import shutil
import numpy as np
import scipy.ndimage
import math
from openpyxl import load_workbook
from openpyxl.utils import (get_column_letter, column_index_from_string, 
                            coordinate_from_string) 
import sympy as smp
import gc

import re
re_pattern_sheet_data = re.compile(r'([A-Za-z]+)([0-9]+[.,]?[0-9]*)')

import matplotlib as mpl
from matplotlib import pyplot as plt, rcParams, ticker
from matplotlib.colors import to_rgb
rcParams['axes.formatter.use_locale'] = True
rcParams.update({'font.family':'serif', 'mathtext.fontset': 'dejavuserif',
                 'font.size': 14, 'axes.titlesize' : 14})

np.set_printoptions(precision=4)

plot_type = namedtuple('plot_type', 'contour contourf contourf_p imshow imshow_p')\
                       ('ax.contour', 'ax.contourf', ('ax.contourf', 'ax.contour'),
                        'ax.imshow', ('ax.imshow', 'ax.contour'))

DIR_WORK = r'C:\Users\ichet\Dropbox\myDocs\St'

IS_SHOW_GRAF = True

print('Розрахунок проводиться в {}.'.format(datetime.strftime(datetime.now(),
                                                          '%d.%m.%Y %H:%M:%S')))    
fig_data = dict(figsize=(8, 6), dpi=100)

# ==============================================================================
def gen_contour(t, plt_type, title, lower_title,
                levels_label_fmt=r'{:.2f}', levels=15, smoothing=10, is_grid=False,
                discolor=False, ax=None, out_fname=None):
    '''Генерує контурні графіки
    args:
    - xlsx_file(basestr) - повна назва XLSX-файлу;
    - sheet_name(basestr) - назва аркуша;
    - data_resc(str, tuple, list) - діапазон клітинок для завантаження даних в 
      фоматі CN:CN, якщо вказано xlsx_file, інакше - масив даних діапазону;
    - plt_type(plot_type, str) - тип графіку;
    - title(basestr) - назва графіку;
    - lower_title(basestr) - нижня назва графіку;
    - levels(int) - кількість рівнів;
    - smoothing(int) - згладжування;
    - is_grid(bool) - при True показує сітку,
    - discolor(bool) - при True знебарвлює графік, при False - відображає
                       вміст графіку в кольорі;
    - ax(Axes) - поточний об'єкт графіку;
    - out_fname(basestr, None) - повна назва файлу для збереження графіку (графік 
      при цьому не відображається). Якщо None - графік відображається.
    
    version: 0.4.0
    
    return: None, list
    '''
    
    # Наперед встановлені поділки осей.
    x = np.array([0, 10, 20, 30, 40, 50])
    y = np.array([0, 10, 20, 30, 40])
    
    t = np.array(t).reshape(5, 6)    # Створення матриці потрібної розмірності.
    
    def extvec(vector, di=1):
        new_vector = []
        previous_value = vector[0]
        
        for value in vector[1:]:
            new_vector += np.linspace(previous_value, value, di, False).tolist()
            previous_value = value
        
        new_vector.append(value)
        
        return np.array(new_vector)

    #def interp(a1, a2, f1, f2, a):
     #   return f1 + (f2 - f1) / (a2 - a1) * (a - a1)

    #def interp_ext_array(xvec, yvec, data_array, interp_func):
     #   new_vector 
    X = extvec(x)
    Y = extvec(y)
    c = np.array([extvec(r) for r in t])
    
    T = np.array([extvec(c) for c in c.transpose()]).transpose()
    
    if smoothing:
        X = scipy.ndimage.zoom(X, smoothing)
        Y = scipy.ndimage.zoom(Y, smoothing)
        T = scipy.ndimage.zoom(T, smoothing)
    
    #raise SystemExit
    
    fig = plt.figure(**fig_data)
    ax = fig.add_subplot(111)
    
    ax.figure.subplots_adjust(left=.1, bottom = .16, top=.92, right=.92)
    
    font = lambda fs: mpl.font_manager.FontProperties(family='Times New Roman', 
      style='normal', size=fs, weight='normal', stretch='normal')
    
    #CS = plt.contour(x, y, t)
    Y = Y[: : -1]
    
    # Побудова графіку.
    if not isinstance(plt_type, tuple): plt_type = (plt_type,)
    for plot_func in plt_type:
        if plot_func == plot_type.contour:
            if discolor or plt_type in (plot_type.contourf_p, plot_type.imshow_p):
                CS = eval(plot_func)(X, Y, T, levels, 
                                     linewidths=1.5,
                                     colors='k') #, nchunk=5
            else:
                CS = eval(plot_func)(X, Y, T, levels, 
                                     linewidths=1.5,
                                     cmap='jet') #, nchunk=5
            #if discolor:
                # Всі ізолінії - чорні.
             #   for cs_line in CS.collections: cs_line.set_color('k')
            
            fmt = dict()
            for lbl in CS.levels:
                fmt[lbl] = (levels_label_fmt.format(lbl)).replace('.', ',')
            ax.clabel(CS, CS.levels, inline=True, fmt=fmt)
            for cont_label in CS.labelTexts:
                # Всі надписи - чорні.
                #if discolor: cont_label.set_color('k')
                
                cont_label.set_fontproperties(font(14))
        else:
            if discolor:
                cs_cmap = plt.cm.get_cmap("gray")
                if plt_type == plot_type.imshow_p:
                    cs_cmap_colors = cs_cmap(range(0, cs_cmap.N))
                    # Збільшує прозорість кольорів.
                    for cmap_color in cs_cmap_colors: cmap_color[3] = 0.9
                    cs_cmap = mpl.colors.ListedColormap(colors=cs_cmap_colors)
            else:
                cs_cmap = 'jet'
            
            if plot_func == plot_type.contourf:
                CS = eval(plot_func)(X, Y, T, levels, cmap=cs_cmap) #, nchunk=5
            elif plot_func == plot_type.imshow:
            # np.flipud(T)
                CS = eval(plot_func)(T, interpolation='bilinear', origin='upper',
                                     extent=(0, 50, 0, 40), cmap=cs_cmap)
    
    #if 
    
    #plt.imshow(T, interpolation='bilinear', origin='upper', extent=(0, 50, 40, 0))
    # Назва графіку.
    title = ax.set_title(title, ha='center', 
                         fontproperties=font(14))
    title.set_position([title.get_position()[0], -0.125])
    
    # Нижня назва графіку.
    lower_title = ax.text(title.get_position()[0], -.15, lower_title, 
      ha='center', va='top', fontproperties=font(14), transform=ax.transAxes)
    
    
    ax.set_xticklabels(['0', '0,1', '0,2', '0,3', '0,4', '0,5'], 
                       fontproperties=font(14))
    ax.set_yticks([0, 10, 20, 30, 40])
    ax.set_yticklabels(['0', '0,1', '0,2', '0,3', '0,4'],
                       fontproperties=font(14))
    # Мітка осі x.
    xl = ax.set_xlabel('a, м', ha='left', 
      va='top', fontproperties=font(14))
    ticklab = ax.xaxis.get_ticklabels()[0]
    trans = ticklab.get_transform()
    ax.xaxis.set_label_coords(52, 0, transform=trans)
    
    yl = ax.set_ylabel('b, м', rotation=0, 
      ha='right', va='bottom',
      fontproperties=font(14))
    ticklab = ax.yaxis.get_ticklabels()[1]
    trans = ticklab.get_transform()
    ax.yaxis.set_label_coords(0, 42, 
      transform=trans)
    
    #CBI = plt.colorbar(CS, orientation='vertical', shrink=1.0)
    
    #plt.axes().axhline(linewidth=1.7, color="red")
    #plt.axes().axvline(linewidth=1.7, color="red")
    
    #for direction in ["xzero", "yzero"]:
     #   print dir(plt.axes().xaxis)
      #  plt.axes().xaxis.set_axisline_style("-|>")
       # plt.axes().xaxis.set_visible(True)

    #for direction in ["left", "right", "bottom", "top"]:
     #   plt.axes().axis[direction].set_visible(False)
    
    #plt.arrow(40, 2, 45, 2, fc='k', ec='k', lw = 1) 
    
    al = 7 # arrow length in points
    arrowprops=dict(#clip_on=False, # plotting outside axes on purpose
      frac=1., # make end arrowhead the whole size of arrow
      headwidth=al, # in points
      facecolor='g')
    kwargs = dict(  
                  xycoords='axes fraction',
                  textcoords='axes fraction',
                  arrowprops= dict(arrowstyle="->"),
                  size=20.
               )
    
    # Розміщення міток підписів осей.
    ax.annotate("",xy=(1.085,0),xytext=(1, 0), **kwargs) # bottom spine arrow
    ax.annotate("",(0,1.085),xytext=(0, 1), **kwargs) # left spin arrow
    
    # Видимість сітки.
    if is_grid:
        ax.grid(b=True, color='#00ff00',
                linestyle='-.',
                linewidth=.5)
    else: ax.grid(False)
    
    #ax.relim()
    #ax.autoscale_view()
    #ax.set_autoscaley_on(True)
    
    
    #ax.set_fontproperties(font)
    
    #for item in ([ax.title, lower_title, ax.xaxis.label, ax.yaxis.label] +
     #        ax.get_xticklabels()+ax.get_yticklabels()):
      #  item.set_fontproperties(font)
    
    if out_fname:
        plt.savefig(out_fname+'.png', dpi=fig_data['dpi'], format='PNG')
        
        # Очищає пам'ять.
        ax.figure.clf()
        plt.close()
        gc.collect()
    else:
        plt.show()
    
def plot_graphs(wb_names, title_parts, cell_starts, dirs):
    '''Побудова контурних графіків
    '''
    
    for i, wb_name in enumerate(wb_names):
        wb = load_workbook(filename=os.path.join(DIR_WORK, 
                             wb_name), read_only=True, data_only=True)
        title_part = title_parts[i]
        print((' Побудова контурних графіків для "{}": '.format(title_part)
               .center(80, '*')))
        
        cell_start = cell_starts[i]
        cell_start_column, cell_start_row = coordinate_from_string(cell_start)
        cell_start_column_index = column_index_from_string(cell_start_column)
        
        target_dir = os.path.join(DIR_WORK, dirs[i])
        try:
            # Спочатку видаляється папка.
            shutil.rmtree(target_dir)
        except FileNotFoundError: pass
        os.mkdir(target_dir)   # Потім створюється папка.
        
        sheets = wb.get_sheet_names()
        sheets_count = len(sheets)
        for i, sheet_name in enumerate(sheets):
            print(u'{}/{} = {:.1%} => WB: {}, WS: {}'.format(i+1, sheets_count,
                                                             (i+1)/sheets_count,
                                                             title_part, sheet_name))
            sheet_name_data = dict(re_pattern_sheet_data.findall(sheet_name))
            if 'l' in sheet_name_data.keys():
                sheet_name_data['L'] = sheet_name_data.pop('l')
            subtitle = (r'd = {d} мм, L = {L} мм, x = {x} см, v = {v} л/(хв$\cdot$м${{}}^2$), '
                        r'$V$ = {V} м${{}}^3$, $I$ = {I} Вт/м${{}}^2$'.format(
                            **sheet_name_data) )
            ws = wb.get_sheet_by_name(sheet_name)
            
            ci = 0  # Індекс стовпця на робочому аркуші.
            while True:
                # прохід по стовпцях робочого аркуша.
                t = []
                cell_start_column_new = get_column_letter(cell_start_column_index+ci)
                cell_start_value_begin = '{}{}'.format(cell_start_column_new,
                                                       cell_start_row+1)
                cell_start_value_end = '{}{}'.format(cell_start_column_new,
                                                     cell_start_row+30)
                for k, c in enumerate(ws[cell_start_value_begin: cell_start_value_end]):
                    cell_value = c[0].value
                    if not k and (cell_value is None):
                        # Припинити читання, якщо перша кл. стовпця - нішо.
                        break
                    t.append(float(cell_value))
                
                if not t:
                    break   # Перехід до наступного аркуша, якщо нічого будувати.
                
                tau = ws['{}{}'.format(cell_start_column_new, cell_start_row)].value
                gen_contour(t, plot_type.imshow_p, r'{} для $\tau$ = {}'.format(
                    title_part, tau), subtitle, out_fname=os.path.join(target_dir, 
                                        '{}-TAU={}'.format(sheet_name, tau)))
                ci += 1

#plot_graphs((r'Змієвик.М1Режим.Циркуляції.Розподіл.тем.на пов.V15I15.xlsx',),
#            ('Змієвик.М1Режим.Циркуляція',),
#            ['B2'], ['contour_ZMIJ_CIRC_M1'])
plot_graphs((r'Гребінка.М1Режим.Циркуляції.Розподіл.тем.на пов.V15I15.xlsx',
             r'Змієвик.М1Режим.Циркуляції.Розподіл.тем.на пов.V15I15.xlsx'),
            ('Гребінка.М1Режим.Циркуляція', 'Змієвик.М1Режим.Циркуляція'),
            ['B1', 'B2'], ['contour_GREB_CIRC_M1', 'contour_ZMIJ_CIRC_M1'])

print(' Роботу завершено! '.center(80, '='))

if IS_SHOW_GRAF:
    plt.show()
    